"""
Export Raw Material Pricing Data to Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import roller_standards as rs

def create_raw_materials_excel():
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="center", vertical="center")
    currency_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def style_cell(ws, row, col, is_currency=False):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = currency_alignment if is_currency else cell_alignment
    
    # ==================== SHEET 1: SUMMARY ====================
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = "CONVEYOR ROLLER - RAW MATERIAL PRICING"
    ws['A1'].font = Font(bold=True, size=16, color="FF6B00")
    ws.merge_cells('A1:D1')
    
    ws['A3'] = "Sheet Index"
    ws['A3'].font = Font(bold=True, size=12)
    
    sheets = [
        ("1. Pipe Specifications", "IS-9295 Pipe diameters and weights per IS-1239"),
        ("2. Shaft Specifications", "Shaft diameters and weights"),
        ("3. Bearing Prices", "Bearing prices by make (China, SKF, FAG, Timken)"),
        ("4. Housing Prices", "Housing prices by pipe-bearing combination"),
        ("5. Seal Prices", "Seal set prices by bearing type"),
        ("6. Circlip Prices", "Circlip prices by shaft diameter"),
        ("7. Rubber Ring Prices", "Impact roller rubber ring prices"),
        ("8. Locking Ring Prices", "Impact roller locking ring prices"),
        ("9. Discount Slabs", "Value-based discount percentages"),
        ("10. Freight Rates", "Distance-based freight rates"),
    ]
    
    for i, (sheet, desc) in enumerate(sheets, start=4):
        ws.cell(row=i, column=1, value=sheet).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    
    # ==================== SHEET 2: PIPE SPECIFICATIONS ====================
    ws_pipe = wb.create_sheet("Pipe Specifications")
    
    ws_pipe['A1'] = "PIPE SPECIFICATIONS (IS-9295 & IS-1239)"
    ws_pipe['A1'].font = Font(bold=True, size=14)
    ws_pipe.merge_cells('A1:F1')
    
    ws_pipe['A3'] = f"Pipe Cost: Rs. {rs.PIPE_COST_PER_KG}/kg"
    ws_pipe['A3'].font = Font(bold=True, color="FF6B00")
    
    headers = ["Pipe OD (mm)", "Type A (Light) kg/m", "Type B (Medium) kg/m", "Type C (Heavy) kg/m"]
    for col, header in enumerate(headers, 1):
        ws_pipe.cell(row=5, column=col, value=header)
    style_header(ws_pipe, 5, len(headers))
    
    row = 6
    for pipe_dia in rs.PIPE_DIAMETERS:
        weights = rs.PIPE_WEIGHT_PER_METER.get(pipe_dia, {})
        ws_pipe.cell(row=row, column=1, value=pipe_dia)
        ws_pipe.cell(row=row, column=2, value=weights.get("A", "-"))
        ws_pipe.cell(row=row, column=3, value=weights.get("B", "-"))
        ws_pipe.cell(row=row, column=4, value=weights.get("C", "-"))
        for col in range(1, 5):
            style_cell(ws_pipe, row, col)
        row += 1
    
    for col in range(1, 5):
        ws_pipe.column_dimensions[get_column_letter(col)].width = 22
    
    # ==================== SHEET 3: SHAFT SPECIFICATIONS ====================
    ws_shaft = wb.create_sheet("Shaft Specifications")
    
    ws_shaft['A1'] = "SHAFT SPECIFICATIONS"
    ws_shaft['A1'].font = Font(bold=True, size=14)
    
    ws_shaft['A3'] = f"Shaft Cost: Rs. {rs.SHAFT_COST_PER_KG}/kg"
    ws_shaft['A3'].font = Font(bold=True, color="FF6B00")
    
    headers = ["Shaft Diameter (mm)", "Weight (kg/m)", "Cost/meter (Rs.)"]
    for col, header in enumerate(headers, 1):
        ws_shaft.cell(row=5, column=col, value=header)
    style_header(ws_shaft, 5, len(headers))
    
    row = 6
    for shaft_dia, weight in rs.SHAFT_WEIGHT_PER_METER.items():
        ws_shaft.cell(row=row, column=1, value=shaft_dia)
        ws_shaft.cell(row=row, column=2, value=weight)
        ws_shaft.cell(row=row, column=3, value=round(weight * rs.SHAFT_COST_PER_KG, 2))
        for col in range(1, 4):
            style_cell(ws_shaft, row, col, col == 3)
        row += 1
    
    for col in range(1, 4):
        ws_shaft.column_dimensions[get_column_letter(col)].width = 20
    
    # ==================== SHEET 4: BEARING PRICES ====================
    ws_bearing = wb.create_sheet("Bearing Prices")
    
    ws_bearing['A1'] = "BEARING PRICES (Rs. per piece)"
    ws_bearing['A1'].font = Font(bold=True, size=14)
    
    headers = ["Shaft (mm)", "Bearing No.", "Bearing OD (mm)", "China", "SKF", "FAG", "Timken"]
    for col, header in enumerate(headers, 1):
        ws_bearing.cell(row=3, column=col, value=header)
    style_header(ws_bearing, 3, len(headers))
    
    row = 4
    for shaft_dia, bearings in rs.BEARING_OPTIONS.items():
        for bearing in bearings:
            ws_bearing.cell(row=row, column=1, value=shaft_dia)
            ws_bearing.cell(row=row, column=2, value=bearing)
            ws_bearing.cell(row=row, column=3, value=rs.BEARING_OD.get(bearing, "-"))
            
            prices = rs.BEARING_COSTS.get(bearing, {})
            ws_bearing.cell(row=row, column=4, value=prices.get("china", "-"))
            ws_bearing.cell(row=row, column=5, value=prices.get("skf", "-"))
            ws_bearing.cell(row=row, column=6, value=prices.get("fag", "-"))
            ws_bearing.cell(row=row, column=7, value=prices.get("timken", "-"))
            
            for col in range(1, 8):
                style_cell(ws_bearing, row, col, col >= 4)
            row += 1
    
    for col in range(1, 8):
        ws_bearing.column_dimensions[get_column_letter(col)].width = 15
    
    # ==================== SHEET 5: HOUSING PRICES ====================
    ws_housing = wb.create_sheet("Housing Prices")
    
    ws_housing['A1'] = "HOUSING PRICES (Rs. per piece)"
    ws_housing['A1'].font = Font(bold=True, size=14)
    
    headers = ["Housing Code", "Housing OD (mm)", "Bearing Bore (mm)", "Price (Rs.)"]
    for col, header in enumerate(headers, 1):
        ws_housing.cell(row=3, column=col, value=header)
    style_header(ws_housing, 3, len(headers))
    
    row = 4
    for housing_code, price in sorted(rs.HOUSING_COSTS.items()):
        parts = housing_code.split("/")
        ws_housing.cell(row=row, column=1, value=housing_code)
        ws_housing.cell(row=row, column=2, value=parts[0])
        ws_housing.cell(row=row, column=3, value=parts[1])
        ws_housing.cell(row=row, column=4, value=price)
        for col in range(1, 5):
            style_cell(ws_housing, row, col, col == 4)
        row += 1
    
    for col in range(1, 5):
        ws_housing.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== SHEET 6: SEAL PRICES ====================
    ws_seal = wb.create_sheet("Seal Prices")
    
    ws_seal['A1'] = "SEAL SET PRICES (Rs. per set)"
    ws_seal['A1'].font = Font(bold=True, size=14)
    
    headers = ["Bearing Number", "Price per Set (Rs.)"]
    for col, header in enumerate(headers, 1):
        ws_seal.cell(row=3, column=col, value=header)
    style_header(ws_seal, 3, len(headers))
    
    row = 4
    for bearing, price in sorted(rs.SEAL_COSTS.items()):
        ws_seal.cell(row=row, column=1, value=bearing)
        ws_seal.cell(row=row, column=2, value=price)
        for col in range(1, 3):
            style_cell(ws_seal, row, col, col == 2)
        row += 1
    
    for col in range(1, 3):
        ws_seal.column_dimensions[get_column_letter(col)].width = 20
    
    # ==================== SHEET 7: CIRCLIP PRICES ====================
    ws_circlip = wb.create_sheet("Circlip Prices")
    
    ws_circlip['A1'] = "CIRCLIP PRICES (Rs. per piece)"
    ws_circlip['A1'].font = Font(bold=True, size=14)
    ws_circlip['A2'] = "Note: 4 circlips per roller"
    ws_circlip['A2'].font = Font(italic=True, color="666666")
    
    headers = ["Shaft Diameter (mm)", "Price per Piece (Rs.)", "Cost per Roller (Rs.)"]
    for col, header in enumerate(headers, 1):
        ws_circlip.cell(row=4, column=col, value=header)
    style_header(ws_circlip, 4, len(headers))
    
    row = 5
    for shaft_dia, price in sorted(rs.CIRCLIP_COSTS.items()):
        ws_circlip.cell(row=row, column=1, value=shaft_dia)
        ws_circlip.cell(row=row, column=2, value=price)
        ws_circlip.cell(row=row, column=3, value=price * 4)
        for col in range(1, 4):
            style_cell(ws_circlip, row, col, col >= 2)
        row += 1
    
    for col in range(1, 4):
        ws_circlip.column_dimensions[get_column_letter(col)].width = 22
    
    # ==================== SHEET 8: RUBBER RING PRICES ====================
    ws_rubber = wb.create_sheet("Rubber Ring Prices")
    
    ws_rubber['A1'] = "RUBBER RING PRICES - Impact Rollers (Rs. per ring)"
    ws_rubber['A1'].font = Font(bold=True, size=14)
    ws_rubber['A2'] = f"Ring Thickness: {rs.RUBBER_RING_THICKNESS}mm | Number of rings = Pipe Length / {rs.RUBBER_RING_THICKNESS}"
    ws_rubber['A2'].font = Font(italic=True, color="666666")
    
    headers = ["Pipe OD (mm)", "Rubber OD (mm)", "Code", "Price per Ring (Rs.)"]
    for col, header in enumerate(headers, 1):
        ws_rubber.cell(row=4, column=col, value=header)
    style_header(ws_rubber, 4, len(headers))
    
    row = 5
    for code, price in sorted(rs.RUBBER_RING_COSTS.items()):
        parts = code.split("/")
        ws_rubber.cell(row=row, column=1, value=parts[0])
        ws_rubber.cell(row=row, column=2, value=parts[1])
        ws_rubber.cell(row=row, column=3, value=code)
        ws_rubber.cell(row=row, column=4, value=price)
        for col in range(1, 5):
            style_cell(ws_rubber, row, col, col == 4)
        row += 1
    
    for col in range(1, 5):
        ws_rubber.column_dimensions[get_column_letter(col)].width = 20
    
    # ==================== SHEET 9: LOCKING RING PRICES ====================
    ws_locking = wb.create_sheet("Locking Ring Prices")
    
    ws_locking['A1'] = "LOCKING RING PRICES - Impact Rollers (Rs. per roller)"
    ws_locking['A1'].font = Font(bold=True, size=14)
    
    headers = ["Pipe OD (mm)", "Price per Roller (Rs.)"]
    for col, header in enumerate(headers, 1):
        ws_locking.cell(row=3, column=col, value=header)
    style_header(ws_locking, 3, len(headers))
    
    row = 4
    for pipe_od, price in sorted(rs.LOCKING_RING_COSTS.items()):
        ws_locking.cell(row=row, column=1, value=pipe_od)
        ws_locking.cell(row=row, column=2, value=price)
        for col in range(1, 3):
            style_cell(ws_locking, row, col, col == 2)
        row += 1
    
    for col in range(1, 3):
        ws_locking.column_dimensions[get_column_letter(col)].width = 22
    
    # ==================== SHEET 10: DISCOUNT SLABS ====================
    ws_discount = wb.create_sheet("Discount Slabs")
    
    ws_discount['A1'] = "VALUE-BASED DISCOUNT SLABS"
    ws_discount['A1'].font = Font(bold=True, size=14)
    
    headers = ["Order Value From (Rs.)", "Order Value To (Rs.)", "Discount %"]
    for col, header in enumerate(headers, 1):
        ws_discount.cell(row=3, column=col, value=header)
    style_header(ws_discount, 3, len(headers))
    
    row = 4
    for min_val, max_val, discount in rs.DISCOUNT_SLABS:
        ws_discount.cell(row=row, column=1, value=f"{min_val:,.0f}")
        ws_discount.cell(row=row, column=2, value="Above" if max_val == float('inf') else f"{max_val:,.0f}")
        ws_discount.cell(row=row, column=3, value=f"{discount}%")
        for col in range(1, 4):
            style_cell(ws_discount, row, col)
        row += 1
    
    for col in range(1, 4):
        ws_discount.column_dimensions[get_column_letter(col)].width = 22
    
    # ==================== SHEET 11: FREIGHT RATES ====================
    ws_freight = wb.create_sheet("Freight Rates")
    
    ws_freight['A1'] = "FREIGHT RATES (Distance-based)"
    ws_freight['A1'].font = Font(bold=True, size=14)
    ws_freight['A2'] = f"Dispatch Location: {rs.DISPATCH_PINCODE} (Gujarat)"
    ws_freight['A2'].font = Font(italic=True, color="666666")
    
    headers = ["Distance From (km)", "Distance To (km)", "Rate (Rs./kg)"]
    for col, header in enumerate(headers, 1):
        ws_freight.cell(row=4, column=col, value=header)
    style_header(ws_freight, 4, len(headers))
    
    row = 5
    for (min_dist, max_dist), rate in rs.FREIGHT_RATES_PER_KG.items():
        ws_freight.cell(row=row, column=1, value=min_dist)
        ws_freight.cell(row=row, column=2, value=f"{max_dist}+" if max_dist == 9999 else max_dist)
        ws_freight.cell(row=row, column=3, value=rate)
        for col in range(1, 4):
            style_cell(ws_freight, row, col, col == 3)
        row += 1
    
    for col in range(1, 4):
        ws_freight.column_dimensions[get_column_letter(col)].width = 20
    
    # Save file
    filename = "/app/backend/raw_materials_pricing.xlsx"
    wb.save(filename)
    print(f"Excel file saved: {filename}")
    return filename

if __name__ == "__main__":
    create_raw_materials_excel()
