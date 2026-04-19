"""Export Routes (Excel/PDF for Quotes, Customers, Products, Cart)"""
from fastapi import APIRouter, HTTPException, Depends, Header
from fastapi.responses import StreamingResponse, Response
from typing import Optional
from datetime import datetime, timezone
from routes import db, get_current_user, get_ist_now, utc_to_ist, IST, SECRET_KEY, ALGORITHM
from jose import jwt, JWTError
from bson import ObjectId
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

@router.get("/quotes/export/excel")
async def export_quotes_excel(
    status: str = None,
    search: str = None,
    token: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    """Export quotes to Excel file. Accepts token as query param or Authorization header."""
    # Validate token from query param OR Authorization header
    current_user = None
    auth_token = token
    
    # Try to get token from Authorization header if not in query
    if not auth_token and authorization:
        if authorization.startswith("Bearer "):
            auth_token = authorization[7:]
    
    if auth_token:
        try:
            payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
            email = payload.get("sub")
            if email:
                user = await db.users.find_one({"email": email})
                if user:
                    user["id"] = str(user["_id"])
                    current_user = user
        except Exception as e:
            logging.error(f"Token validation error: {e}")
            raise HTTPException(status_code=401, detail="Invalid token")
    
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        # Build query
        query = {}
        if current_user.get("role") != "admin":
            query["customer_email"] = current_user.get("email")
        if status and status != "all":
            query["status"] = status
        if search:
            query["$or"] = [
                {"quote_number": {"$regex": search, "$options": "i"}},
                {"customer_name": {"$regex": search, "$options": "i"}},
                {"customer_company": {"$regex": search, "$options": "i"}}
            ]
        
        quotes = await db.quotes.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotes"
        
        # Headers
        headers = ["Quote Number", "Customer", "Company", "Status", "Products", "Subtotal", "Discount", "Packing", "Freight", "Total", "Created Date"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="960018", end_color="960018", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, quote in enumerate(quotes, 2):
            ws.cell(row=row, column=1, value=quote.get("quote_number", "N/A"))
            ws.cell(row=row, column=2, value=quote.get("customer_name", "N/A"))
            ws.cell(row=row, column=3, value=quote.get("customer_company", "N/A"))
            ws.cell(row=row, column=4, value=quote.get("status", "N/A"))
            ws.cell(row=row, column=5, value=len(quote.get("products", [])))
            ws.cell(row=row, column=6, value=quote.get("subtotal", 0))
            ws.cell(row=row, column=7, value=quote.get("total_discount", 0))
            ws.cell(row=row, column=8, value=quote.get("packing_charges", 0))
            ws.cell(row=row, column=9, value=quote.get("shipping_cost", 0))
            ws.cell(row=row, column=10, value=quote.get("total_price", 0))
            created = quote.get("created_at")
            if created:
                ws.cell(row=row, column=11, value=created.strftime("%Y-%m-%d %H:%M") if hasattr(created, 'strftime') else str(created)[:16])
        
        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Quotes_Export_{get_ist_now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logging.error(f"Quote Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/quotes/export/pdf")
async def export_quotes_pdf(
    status: str = None,
    token: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    """Export quotes to PDF file. Accepts token as query param or Authorization header."""
    # Validate token from query param OR Authorization header
    current_user = None
    auth_token = token
    
    # Try to get token from Authorization header if not in query
    if not auth_token and authorization:
        if authorization.startswith("Bearer "):
            auth_token = authorization[7:]
    
    if auth_token:
        try:
            payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
            email = payload.get("sub")
            if email:
                user = await db.users.find_one({"email": email})
                if user:
                    user["id"] = str(user["_id"])
                    current_user = user
        except Exception as e:
            logging.error(f"Token validation error: {e}")
            raise HTTPException(status_code=401, detail="Invalid token")
    
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        query = {}
        if current_user.get("role") != "admin":
            query["customer_email"] = current_user.get("email")
        if status:
            query["status"] = status
        
        quotes = await db.quotes.find(query, {
            "quote_number": 1, "customer_name": 1, "customer_company": 1,
            "total_price": 1, "status": 1, "created_at": 1, "items": 1
        }).sort("created_at", -1).limit(500).to_list(500)
        
        # Generate PDF HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Quotes Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #960018; border-bottom: 2px solid #960018; padding-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background-color: #960018; color: white; padding: 12px; text-align: left; }}
                td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .status {{ padding: 4px 8px; border-radius: 4px; font-size: 12px; }}
                .approved {{ background-color: #4CAF50; color: white; }}
                .pending {{ background-color: #FF9800; color: white; }}
                .rejected {{ background-color: #f44336; color: white; }}
                .footer {{ margin-top: 30px; text-align: center; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <h1>Quotes Export</h1>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <table>
                <thead>
                    <tr>
                        <th>Quote #</th>
                        <th>Customer</th>
                        <th>Items</th>
                        <th>Total</th>
                        <th>Status</th>
                        <th>Date</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for quote in quotes:
            status_class = quote.get('status', 'pending').lower().replace('rfq_', '')
            created = quote.get('created_at', datetime.now())
            if isinstance(created, str):
                created = datetime.fromisoformat(created.replace('Z', '+00:00'))
            
            html_content += f"""
                    <tr>
                        <td>{quote.get('quote_number', 'N/A')}</td>
                        <td>{quote.get('customer_name', 'N/A')}</td>
                        <td>{len(quote.get('products', []))}</td>
                        <td>Rs. {quote.get('total_price', 0):,.2f}</td>
                        <td><span class="status {status_class}">{quote.get('status', 'N/A').upper()}</span></td>
                        <td>{created.strftime('%Y-%m-%d')}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
            <div class="footer">
                <p>Convero - Belt Conveyor Roller Solutions</p>
            </div>
        </body>
        </html>
        """
        
        # Generate PDF using weasyprint
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = f"quotes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logging.error(f"Quote PDF export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/customers/export/excel")
async def export_customers_excel(
    search: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Export customers to Excel file - Admin only"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        query = {"role": "customer"}
        if search:
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"company": {"$regex": search, "$options": "i"}},
                {"customer_code": {"$regex": search, "$options": "i"}}
            ]
        
        customers = await db.users.find(query, {
            "_id": 0, "password": 0, "push_token": 0
        }).limit(500).to_list(500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        headers = ["Customer Code", "Name", "Email", "Company", "Phone", "City", "State", "GST Number", "Created Date"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="960018", end_color="960018", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row, customer in enumerate(customers, 2):
            ws.cell(row=row, column=1, value=customer.get("customer_code", "N/A"))
            ws.cell(row=row, column=2, value=customer.get("name", "N/A"))
            ws.cell(row=row, column=3, value=customer.get("email", "N/A"))
            ws.cell(row=row, column=4, value=customer.get("company", "N/A"))
            ws.cell(row=row, column=5, value=customer.get("phone", "N/A"))
            ws.cell(row=row, column=6, value=customer.get("city", "N/A"))
            ws.cell(row=row, column=7, value=customer.get("state", "N/A"))
            ws.cell(row=row, column=8, value=customer.get("gst_number", "N/A"))
            created = customer.get("created_at")
            if created:
                ws.cell(row=row, column=9, value=created.strftime("%Y-%m-%d") if hasattr(created, 'strftime') else str(created)[:10])
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Customers_Export_{get_ist_now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logging.error(f"Customer Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/customers/export/pdf")
async def export_customers_pdf(
    search: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Export customers to PDF file - Admin only"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        query = {}
        if search:
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"company": {"$regex": search, "$options": "i"}}
            ]
        
        customers = await db.users.find({**query, "role": "customer"}, {
            "_id": 0, "password": 0, "push_token": 0
        }).sort("created_at", -1).limit(500).to_list(500)
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Customers Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #960018; border-bottom: 2px solid #960018; padding-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background-color: #960018; color: white; padding: 12px; text-align: left; }}
                td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .footer {{ margin-top: 30px; text-align: center; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <h1>Customer List</h1>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(customers)} customers</p>
            <table>
                <thead>
                    <tr>
                        <th>Customer Code</th>
                        <th>Name</th>
                        <th>Email</th>
                        <th>Phone</th>
                        <th>Company</th>
                        <th>Joined</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for customer in customers:
            created = customer.get('created_at', datetime.now())
            if isinstance(created, str):
                created = datetime.fromisoformat(created.replace('Z', '+00:00'))
            
            html_content += f"""
                    <tr>
                        <td>{customer.get('customer_code', 'N/A')}</td>
                        <td>{customer.get('name', 'N/A')}</td>
                        <td>{customer.get('email', 'N/A')}</td>
                        <td>{customer.get('phone', 'N/A')}</td>
                        <td>{customer.get('company', 'N/A')}</td>
                        <td>{created.strftime('%Y-%m-%d')}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
            <div class="footer">
                <p>Convero - Belt Conveyor Roller Solutions</p>
            </div>
        </body>
        </html>
        """
        
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = f"customers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logging.error(f"Customer PDF export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/products/export/excel")
async def export_products_excel(
    search: str = None,
    roller_type: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Export product catalog to Excel file"""
    try:
        query = {}
        if search:
            query["$or"] = [
                {"product_code": {"$regex": search, "$options": "i"}},
                {"product_name": {"$regex": search, "$options": "i"}}
            ]
        if roller_type:
            query["roller_type"] = roller_type
        
        products = await db.products.find(query, {"_id": 0}).to_list(5000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        headers = ["Product Code", "Product Name", "Roller Type", "Pipe OD", "Shaft Dia", "Bearing", "Face Length", "Weight", "Unit Price"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="960018", end_color="960018", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product.get("product_code", "N/A"))
            ws.cell(row=row, column=2, value=product.get("product_name", "N/A"))
            ws.cell(row=row, column=3, value=product.get("roller_type", "N/A"))
            ws.cell(row=row, column=4, value=product.get("pipe_od", "N/A"))
            ws.cell(row=row, column=5, value=product.get("shaft_dia", "N/A"))
            ws.cell(row=row, column=6, value=product.get("bearing", "N/A"))
            ws.cell(row=row, column=7, value=product.get("face_length", "N/A"))
            ws.cell(row=row, column=8, value=product.get("weight", 0))
            ws.cell(row=row, column=9, value=product.get("unit_price", 0))
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Products_Export_{get_ist_now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logging.error(f"Product Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/products/export/pdf")
async def export_products_pdf(
    search: str = None,
    roller_type: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Export product catalog to PDF file"""
    try:
        query = {}
        if search:
            query["$or"] = [
                {"product_code": {"$regex": search, "$options": "i"}},
                {"product_name": {"$regex": search, "$options": "i"}}
            ]
        if roller_type:
            query["roller_type"] = roller_type
        
        products = await db.products.find(query, {
            "name": 1, "roller_type": 1, "pipe_od": 1, "shaft_dia": 1,
            "bearing_type": 1, "price": 1, "created_at": 1
        }).limit(200).to_list(200)
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Product Catalog</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #960018; border-bottom: 2px solid #960018; padding-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 11px; }}
                th {{ background-color: #960018; color: white; padding: 8px; text-align: left; }}
                td {{ padding: 6px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .footer {{ margin-top: 30px; text-align: center; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <h1>Product Catalog</h1>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(products)} products</p>
            <table>
                <thead>
                    <tr>
                        <th>Code</th>
                        <th>Name</th>
                        <th>Type</th>
                        <th>Pipe Dia</th>
                        <th>Shaft</th>
                        <th>Length</th>
                        <th>Weight</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for product in products:
            html_content += f"""
                    <tr>
                        <td>{product.get('product_code', 'N/A')}</td>
                        <td>{product.get('product_name', 'N/A')}</td>
                        <td>{product.get('roller_type', 'N/A')}</td>
                        <td>{product.get('pipe_diameter', 'N/A')}mm</td>
                        <td>{product.get('shaft_diameter', 'N/A')}mm</td>
                        <td>{product.get('roller_length', 'N/A')}mm</td>
                        <td>{product.get('total_weight', 0):.2f}kg</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
            <div class="footer">
                <p>Convero - Belt Conveyor Roller Solutions</p>
            </div>
        </body>
        </html>
        """
        
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logging.error(f"Product PDF export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/cart/export/excel")
async def export_cart_excel(
    current_user: dict = Depends(get_current_user)
):
    """Export cart contents to Excel file"""
    try:
        user_id = current_user.get("user_id")
        cart = await db.carts.find_one({"user_id": user_id})
        
        if not cart or not cart.get("items"):
            raise HTTPException(status_code=404, detail="Cart is empty")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Cart"
        
        headers = ["Product Code", "Product Name", "Roller Type", "Specifications", "Quantity", "Unit Price", "Total Price", "Weight"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="960018", end_color="960018", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        total_value = 0
        total_weight = 0
        for row, item in enumerate(cart.get("items", []), 2):
            specs = item.get("specifications", {})
            spec_str = f"Pipe: {specs.get('pipe_od', 'N/A')}, Shaft: {specs.get('shaft_dia', 'N/A')}, Bearing: {specs.get('bearing', 'N/A')}"
            
            ws.cell(row=row, column=1, value=item.get("product_code", "N/A"))
            ws.cell(row=row, column=2, value=item.get("product_name", "N/A"))
            ws.cell(row=row, column=3, value=item.get("roller_type", "N/A"))
            ws.cell(row=row, column=4, value=spec_str)
            ws.cell(row=row, column=5, value=item.get("quantity", 0))
            ws.cell(row=row, column=6, value=item.get("unit_price", 0))
            ws.cell(row=row, column=7, value=item.get("total_price", 0))
            ws.cell(row=row, column=8, value=item.get("weight", 0))
            
            total_value += item.get("total_price", 0)
            total_weight += item.get("weight", 0) * item.get("quantity", 0)
        
        # Add totals row
        last_row = len(cart.get("items", [])) + 2
        ws.cell(row=last_row, column=6, value="TOTAL:")
        ws.cell(row=last_row, column=7, value=total_value)
        ws.cell(row=last_row, column=8, value=total_weight)
        ws.cell(row=last_row, column=6).font = Font(bold=True)
        ws.cell(row=last_row, column=7).font = Font(bold=True)
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Cart_Export_{get_ist_now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Cart Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ============= COMPANY DETAILS FOR PDF =============
COMPANY_INFO = {
    "name": "Convero Engineering",
    "address": "Pune, Maharashtra, India",
    "email": "info@convero.in",
    "phone": "",
    "gst": "",
    "bank_name": "",
    "bank_account": "",
    "bank_ifsc": "",
    "bank_branch": "",
}

# ============= ORDERS EXPORT =============

@router.get("/orders/export/excel")
async def export_orders_excel(
    stage: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    try:
        query = {}
        if stage:
            query["stage"] = stage

        orders = await db.sales_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Orders"

        headers = ["SO Number", "Quote Ref", "Customer", "Company", "Stage", "Total Price", "Paid", "Balance Due", "Payment Status", "PI Number", "Invoice Number", "Created"]
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, order in enumerate(orders, 2):
            ws.cell(row=row_idx, column=1, value=order.get("so_number"))
            ws.cell(row=row_idx, column=2, value=order.get("quote_number"))
            ws.cell(row=row_idx, column=3, value=order.get("customer_name"))
            ws.cell(row=row_idx, column=4, value=order.get("customer_company"))
            ws.cell(row=row_idx, column=5, value=order.get("stage"))
            ws.cell(row=row_idx, column=6, value=round(order.get("total_price", 0), 2))
            ws.cell(row=row_idx, column=7, value=round(order.get("total_paid", 0), 2))
            ws.cell(row=row_idx, column=8, value=round(order.get("balance_due", 0), 2))
            ws.cell(row=row_idx, column=9, value=order.get("payment_status"))
            ws.cell(row=row_idx, column=10, value=order.get("proforma_invoice"))
            ws.cell(row=row_idx, column=11, value=order.get("tax_invoice"))
            ws.cell(row=row_idx, column=12, value=str(order.get("created_at", ""))[:10])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Sales_Orders.xlsx"}
        )
    except Exception as e:
        logging.error(f"Orders export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/orders/export/pdf")
async def export_orders_pdf(
    stage: Optional[str] = None,
    token: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    auth_token = token
    if not auth_token and authorization and authorization.startswith("Bearer "):
        auth_token = authorization[7:]
    if not auth_token:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    query = {}
    if stage:
        query["stage"] = stage

    orders = await db.sales_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

    html = f"""<html><head><style>
    body {{ font-family: Arial; margin: 20px; }}
    h1 {{ color: #0F172A; border-bottom: 3px solid #C5964A; padding-bottom: 8px; }}
    .company {{ color: #64748B; font-size: 12px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
    th {{ background: #0F172A; color: white; padding: 10px; text-align: left; font-size: 11px; }}
    td {{ padding: 8px; border-bottom: 1px solid #E2E8F0; font-size: 11px; }}
    .stage {{ padding: 3px 8px; border-radius: 4px; font-weight: bold; font-size: 10px; }}
    .confirmed {{ background: #DBEAFE; color: #1E40AF; }}
    .in_production {{ background: #FEF3C7; color: #92400E; }}
    .ready {{ background: #EDE9FE; color: #5B21B6; }}
    .dispatched {{ background: #FEF3C7; color: #92400E; }}
    .delivered {{ background: #D1FAE5; color: #065F46; }}
    .paid {{ color: #10B981; font-weight: bold; }}
    .partial {{ color: #F59E0B; font-weight: bold; }}
    .unpaid {{ color: #EF4444; font-weight: bold; }}
    </style></head><body>
    <h1>Sales Orders</h1>
    <p class="company">{COMPANY_INFO['name']} | {COMPANY_INFO['email']}</p>
    <table><tr><th>SO #</th><th>Customer</th><th>Stage</th><th>Total</th><th>Paid</th><th>Due</th><th>Payment</th><th>Invoice</th><th>Date</th></tr>"""

    for o in orders:
        stage_cls = o.get("stage", "").replace(" ", "_")
        pay_cls = o.get("payment_status", "unpaid")
        html += f"""<tr>
        <td><b>{o.get('so_number','')}</b></td>
        <td>{o.get('customer_name','')}</td>
        <td><span class="stage {stage_cls}">{o.get('stage','')}</span></td>
        <td>Rs.{o.get('total_price',0):,.2f}</td>
        <td>Rs.{o.get('total_paid',0):,.2f}</td>
        <td>Rs.{o.get('balance_due',0):,.2f}</td>
        <td><span class="{pay_cls}">{pay_cls}</span></td>
        <td>{o.get('tax_invoice','') or o.get('proforma_invoice','') or '-'}</td>
        <td>{str(o.get('created_at',''))[:10]}</td>
        </tr>"""

    html += "</table></body></html>"

    output = io.BytesIO(html.encode())
    output.seek(0)
    return StreamingResponse(output, media_type="text/html",
                           headers={"Content-Disposition": "attachment; filename=Sales_Orders.html"})


# ============= CRM LEADS EXPORT =============

@router.get("/crm/leads/export/excel")
async def export_leads_excel(current_user: dict = Depends(get_current_user)):
    try:
        leads = await db.leads.find({}, {"_id": 0}).sort("updated_at", -1).to_list(500)

        wb = Workbook()
        ws = wb.active
        ws.title = "CRM Leads"

        headers = ["Name", "Company", "Email", "Phone", "Stage", "Source", "Product Interest", "Est. Value", "Created", "Updated"]
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, lead in enumerate(leads, 2):
            ws.cell(row=row_idx, column=1, value=lead.get("name"))
            ws.cell(row=row_idx, column=2, value=lead.get("company"))
            ws.cell(row=row_idx, column=3, value=lead.get("email"))
            ws.cell(row=row_idx, column=4, value=lead.get("phone"))
            ws.cell(row=row_idx, column=5, value=lead.get("stage"))
            ws.cell(row=row_idx, column=6, value=lead.get("source"))
            ws.cell(row=row_idx, column=7, value=lead.get("product_interest"))
            ws.cell(row=row_idx, column=8, value=lead.get("estimated_value"))
            ws.cell(row=row_idx, column=9, value=str(lead.get("created_at", ""))[:10])
            ws.cell(row=row_idx, column=10, value=str(lead.get("updated_at", ""))[:10])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=CRM_Leads.xlsx"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/crm/followups/export/excel")
async def export_followups_excel(current_user: dict = Depends(get_current_user)):
    try:
        followups = await db.followups.find({}, {"_id": 0}).sort("due_date", 1).to_list(500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Follow-ups"

        headers = ["Lead ID", "Type", "Due Date", "Note", "Completed", "Created By"]
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, fu in enumerate(followups, 2):
            lead = await db.leads.find_one({"id": fu.get("lead_id")}, {"_id": 0, "name": 1})
            ws.cell(row=row_idx, column=1, value=lead.get("name") if lead else fu.get("lead_id"))
            ws.cell(row=row_idx, column=2, value=fu.get("follow_up_type"))
            ws.cell(row=row_idx, column=3, value=str(fu.get("due_date", ""))[:10])
            ws.cell(row=row_idx, column=4, value=fu.get("note"))
            ws.cell(row=row_idx, column=5, value="Yes" if fu.get("completed") else "No")
            ws.cell(row=row_idx, column=6, value=fu.get("created_by"))

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=CRM_Followups.xlsx"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============= INVOICES EXPORT =============

@router.get("/invoices/export/excel")
async def export_invoices_excel(
    invoice_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    try:
        query = {}
        if invoice_type:
            query["invoice_type"] = invoice_type

        invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"

        headers = ["Invoice #", "Type", "SO #", "Customer", "Taxable Amount", "CGST", "SGST", "Total with GST", "Created"]
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, inv in enumerate(invoices, 2):
            ws.cell(row=row_idx, column=1, value=inv.get("invoice_number"))
            ws.cell(row=row_idx, column=2, value=inv.get("invoice_type", "").upper())
            ws.cell(row=row_idx, column=3, value=inv.get("so_number"))
            ws.cell(row=row_idx, column=4, value=inv.get("customer_name"))
            ws.cell(row=row_idx, column=5, value=round(inv.get("taxable_amount", inv.get("total_price", 0)), 2))
            ws.cell(row=row_idx, column=6, value=round(inv.get("cgst_amount", 0), 2))
            ws.cell(row=row_idx, column=7, value=round(inv.get("sgst_amount", 0), 2))
            ws.cell(row=row_idx, column=8, value=round(inv.get("total_with_gst", inv.get("total_price", 0)), 2))
            ws.cell(row=row_idx, column=9, value=str(inv.get("created_at", ""))[:10])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Invoices.xlsx"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==============================================================================
# Generic PDF builder + missing PDF/Excel endpoints (Leads, Followups, Invoices,
# Store Stock/POs/Suppliers, Work Orders, WIP QC)
# ==============================================================================

def _build_table_pdf(title: str, headers: list, rows: list, col_widths: list = None, orientation: str = "L") -> bytes:
    """Create a generic table PDF using fpdf2. Returns bytes."""
    from fpdf import FPDF
    pdf = FPDF(orientation=orientation, unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, title, ln=1, align="L")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", ln=1)
    pdf.ln(2)

    usable_w = (pdf.w - 20) if orientation == "P" else (pdf.w - 20)  # 297-20 for landscape A4
    if not col_widths:
        col_widths = [usable_w / max(1, len(headers))] * len(headers)

    # Header row
    pdf.set_fill_color(15, 23, 42)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for h, w in zip(headers, col_widths):
        pdf.cell(w, 7, str(h)[:40], border=1, align="C", fill=True)
    pdf.ln()

    # Data rows
    pdf.set_text_color(15, 23, 42)
    pdf.set_font("Helvetica", "", 8)
    fill = False
    for row in rows:
        if pdf.get_y() > pdf.h - 20:
            pdf.add_page()
            pdf.set_fill_color(15, 23, 42)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 9)
            for h, w in zip(headers, col_widths):
                pdf.cell(w, 7, str(h)[:40], border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_text_color(15, 23, 42)
            pdf.set_font("Helvetica", "", 8)
        if fill:
            pdf.set_fill_color(248, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)
        for v, w in zip(row, col_widths):
            s = "" if v is None else str(v)
            # Strip characters outside latin-1 range (Helvetica core font limitation)
            s = s.replace("—", "-").replace("–", "-").replace("×", "x").replace("₹", "Rs.")
            try:
                s.encode("latin-1")
            except UnicodeEncodeError:
                s = s.encode("latin-1", "replace").decode("latin-1")
            pdf.cell(w, 6, s[:60], border=1, align="L", fill=True)
        pdf.ln()
        fill = not fill

    out = pdf.output(dest="S")
    return bytes(out) if isinstance(out, (bytes, bytearray)) else out.encode("latin-1") if isinstance(out, str) else bytes(out)


def _build_table_excel(sheet_title: str, headers: list, rows: list) -> bytes:
    """Generic Excel builder with navy header."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:30]
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for r_idx, row in enumerate(rows, 2):
        for c_idx, v in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _pdf_response(data: bytes, filename: str):
    return Response(content=data, media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


def _xlsx_response(data: bytes, filename: str):
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ---------- CRM Leads PDF ----------
@router.get("/crm/leads/export/pdf")
async def export_leads_pdf(current_user: dict = Depends(get_current_user)):
    leads = await db.leads.find({}, {"_id": 0}).sort("updated_at", -1).to_list(500)
    headers = ["Name", "Company", "Phone", "Stage", "Source", "Product", "Est. Value", "Updated"]
    rows = [[l.get("name"), l.get("company"), l.get("phone"), l.get("stage"),
             l.get("source"), l.get("product_interest"), l.get("estimated_value"),
             str(l.get("updated_at", ""))[:10]] for l in leads]
    return _pdf_response(_build_table_pdf("CRM Leads", headers, rows), "CRM_Leads.pdf")


# ---------- CRM Followups PDF ----------
@router.get("/crm/followups/export/pdf")
async def export_followups_pdf(current_user: dict = Depends(get_current_user)):
    followups = await db.followups.find({}, {"_id": 0}).sort("due_date", 1).to_list(500)
    rows = []
    for fu in followups:
        lead = await db.leads.find_one({"id": fu.get("lead_id")}, {"_id": 0, "name": 1})
        rows.append([lead.get("name") if lead else fu.get("lead_id"),
                     fu.get("follow_up_type"), str(fu.get("due_date", ""))[:10],
                     (fu.get("note") or "")[:50], "Yes" if fu.get("completed") else "No",
                     fu.get("created_by")])
    return _pdf_response(
        _build_table_pdf("Follow-ups", ["Lead", "Type", "Due Date", "Note", "Done", "Created By"], rows),
        "CRM_Followups.pdf")


# ---------- Invoices PDF ----------
@router.get("/invoices/export/pdf")
async def export_invoices_pdf(
    invoice_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    q = {}
    if invoice_type: q["invoice_type"] = invoice_type
    invs = await db.invoices.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    headers = ["Invoice #", "Type", "SO #", "Customer", "GSTIN", "Subtotal", "GST", "Total", "Date"]
    rows = [[i.get("invoice_number"), i.get("invoice_type"), i.get("so_number"),
             i.get("customer_name"), i.get("customer_gstin"),
             round(i.get("subtotal", 0), 2), round(i.get("gst_amount", 0), 2),
             round(i.get("total_with_gst", i.get("total_price", 0)), 2),
             str(i.get("created_at", ""))[:10]] for i in invs]
    return _pdf_response(_build_table_pdf("Invoices", headers, rows), "Invoices.pdf")


# ---------- Store: Stock PDF ----------
@router.get("/store/export/stock/pdf")
async def export_stock_pdf(current_user: dict = Depends(get_current_user)):
    items = await db.stock_items.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    headers = ["Name", "Category", "Unit", "On Hand", "Reserved", "Available", "Reorder At", "Avg Rate"]
    rows = [[i.get("name"), i.get("category"), i.get("unit_purchase"),
             i.get("on_hand", 0), i.get("reserved", 0),
             (i.get("on_hand", 0) - i.get("reserved", 0)),
             i.get("reorder_level"), i.get("avg_rate")] for i in items]
    return _pdf_response(_build_table_pdf("Stock Items", headers, rows), "Stock.pdf")


# ---------- Store: POs PDF ----------
@router.get("/store/export/purchase-orders/pdf")
async def export_pos_pdf(current_user: dict = Depends(get_current_user)):
    pos = await db.purchase_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    headers = ["PO #", "Supplier", "Status", "Total", "PO Date", "Expected", "Items"]
    rows = [[p.get("po_number"), p.get("supplier_name"), p.get("status"),
             round(p.get("total_amount", 0), 2), str(p.get("po_date", ""))[:10],
             str(p.get("expected_date", ""))[:10], len(p.get("items") or [])] for p in pos]
    return _pdf_response(_build_table_pdf("Purchase Orders", headers, rows), "POs.pdf")


# ---------- Store: Suppliers PDF ----------
@router.get("/store/export/suppliers/pdf")
async def export_suppliers_pdf(current_user: dict = Depends(get_current_user)):
    sups = await db.suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    headers = ["Name", "GSTIN", "Phone", "Email", "Address", "Payment Terms"]
    rows = [[s.get("name"), s.get("gstin"), s.get("contact_phone"),
             s.get("contact_email"), (s.get("address") or "")[:40],
             s.get("payment_terms")] for s in sups]
    return _pdf_response(_build_table_pdf("Suppliers", headers, rows), "Suppliers.pdf")


# ---------- Work Orders EXCEL + PDF ----------
@router.get("/work-orders/export/excel")
async def export_work_orders_excel(current_user: dict = Depends(get_current_user)):
    wos = await db.work_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    headers = ["WO #", "SO #", "Customer", "Stage", "Items", "Delivery Date", "Created"]
    rows = [[w.get("wo_number"), w.get("so_number"), w.get("customer_name"),
             w.get("stage"), len(w.get("items") or []),
             str(w.get("delivery_date", ""))[:10],
             str(w.get("created_at", ""))[:10]] for w in wos]
    return _xlsx_response(_build_table_excel("Work Orders", headers, rows), "Work_Orders.xlsx")


@router.get("/work-orders/export/pdf")
async def export_work_orders_pdf(current_user: dict = Depends(get_current_user)):
    wos = await db.work_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    headers = ["WO #", "SO #", "Customer", "Stage", "Items", "Delivery", "Created"]
    rows = [[w.get("wo_number"), w.get("so_number"), w.get("customer_name"),
             w.get("stage"), len(w.get("items") or []),
             str(w.get("delivery_date", ""))[:10],
             str(w.get("created_at", ""))[:10]] for w in wos]
    return _pdf_response(_build_table_pdf("Work Orders", headers, rows), "Work_Orders.pdf")


# ---------- WIP QC Overview EXCEL + PDF ----------
async def _wip_qc_rows():
    wos = await db.work_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    subs = await db.sub_work_orders.find({}, {"_id": 0}).to_list(3000)
    sub_by = {}
    for s in subs:
        sub_by.setdefault(s.get("parent_wo_id"), []).append(s)
    rows = []
    for w in wos:
        items = w.get("items") or []
        if not any("pulley" not in (it.get("product_name") or "").lower() for it in items):
            continue
        ss = sub_by.get(w.get("id")) or []
        pipe = next((x for x in ss if x.get("type") == "pipe"), None)
        shaft = next((x for x in ss if x.get("type") == "shaft"), None)
        def _st(sub):
            if not sub: return ("—", "—", "—")
            wip = sub.get("wip_qc") or {}
            pc = sum(int(i.get("pass_count") or 0) for i in wip.get("items") or [])
            fc = sum(int(i.get("fail_count") or 0) for i in wip.get("items") or [])
            return (wip.get("status") or "pending", f"{pc}/{fc}", wip.get("inspected_by") or "—")
        ps, pcnt, pby = _st(pipe)
        ss2, scnt, sby = _st(shaft)
        rows.append([w.get("wo_number"), w.get("so_number"), w.get("customer_name"),
                     w.get("stage"), ps, pcnt, pby, ss2, scnt, sby])
    return rows

@router.get("/wip-qc/export/excel")
async def export_wip_qc_excel(current_user: dict = Depends(get_current_user)):
    headers = ["WO #", "SO #", "Customer", "Stage",
               "Pipe Status", "Pipe P/F", "Pipe Inspector",
               "Shaft Status", "Shaft P/F", "Shaft Inspector"]
    rows = await _wip_qc_rows()
    return _xlsx_response(_build_table_excel("WIP QC Overview", headers, rows), "WIP_QC_Overview.xlsx")


@router.get("/wip-qc/export/pdf")
async def export_wip_qc_pdf(current_user: dict = Depends(get_current_user)):
    headers = ["WO #", "SO #", "Customer", "Stage",
               "Pipe", "P/F", "Inspector",
               "Shaft", "P/F", "Inspector"]
    rows = await _wip_qc_rows()
    return _pdf_response(_build_table_pdf("WIP QC Overview", headers, rows), "WIP_QC_Overview.pdf")

