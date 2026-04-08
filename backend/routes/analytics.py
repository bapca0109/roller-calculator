"""Analytics & Dashboard Routes"""
from fastapi import APIRouter, HTTPException, Depends, Header
from fastapi.responses import StreamingResponse
from typing import Optional, Dict, Any
from datetime import datetime, timedelta, timezone
from routes import db, get_current_user, get_ist_now, utc_to_ist, IST
from bson import ObjectId
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

@router.get("/analytics/dashboard")
async def get_dashboard_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive dashboard analytics (admin only)"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Parse date filters
        date_filter = {}
        if start_date and end_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                date_filter = {"created_at": {"$gte": start_dt, "$lte": end_dt}}
            except ValueError:
                pass
        
        # Get current date info
        now = get_ist_now()
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        
        # Get financial year dates
        if now.month >= 4:
            fy_start = now.replace(month=4, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            fy_start = now.replace(year=now.year-1, month=4, day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Total quotes count
        quotes_filter = {**date_filter} if date_filter else {}
        total_quotes = await db.quotes.count_documents(quotes_filter)
        
        # Approved quotes count
        approved_filter = {"status": "approved", **date_filter} if date_filter else {"status": "approved"}
        approved_quotes = await db.quotes.count_documents(approved_filter)
        
        # Pending RFQs count
        pending_filter = {"status": {"$ne": "approved"}, **date_filter} if date_filter else {"status": {"$ne": "approved"}}
        pending_rfqs = await db.quotes.count_documents(pending_filter)
        
        # Total customers
        customer_filter = {"role": "customer", **date_filter} if date_filter else {"role": "customer"}
        total_customers = await db.users.count_documents({"role": "customer"})
        
        # New customers this month
        new_customers_this_month = await db.users.count_documents({
            "role": "customer",
            "created_at": {"$gte": current_month_start.replace(tzinfo=None)}
        })
        
        # Calculate total revenue from approved quotes
        revenue_match = {"status": "approved", **date_filter} if date_filter else {"status": "approved"}
        revenue_pipeline = [
            {"$match": revenue_match},
            {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
        ]
        revenue_result = await db.quotes.aggregate(revenue_pipeline).to_list(1)
        total_revenue = revenue_result[0]["total"] if revenue_result else 0
        
        # Revenue this month
        monthly_revenue_pipeline = [
            {"$match": {
                "status": "approved",
                "created_at": {"$gte": current_month_start.replace(tzinfo=None)}
            }},
            {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
        ]
        monthly_revenue_result = await db.quotes.aggregate(monthly_revenue_pipeline).to_list(1)
        monthly_revenue = monthly_revenue_result[0]["total"] if monthly_revenue_result else 0
        
        # Revenue last month for comparison
        last_month_revenue_pipeline = [
            {"$match": {
                "status": "approved",
                "created_at": {
                    "$gte": last_month_start.replace(tzinfo=None),
                    "$lt": current_month_start.replace(tzinfo=None)
                }
            }},
            {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
        ]
        last_month_revenue_result = await db.quotes.aggregate(last_month_revenue_pipeline).to_list(1)
        last_month_revenue = last_month_revenue_result[0]["total"] if last_month_revenue_result else 0
        
        # Calculate revenue growth percentage
        if last_month_revenue > 0:
            revenue_growth = ((monthly_revenue - last_month_revenue) / last_month_revenue) * 100
        else:
            revenue_growth = 100 if monthly_revenue > 0 else 0
        
        # Average quote value
        avg_quote_value = total_revenue / approved_quotes if approved_quotes > 0 else 0
        
        # Conversion rate (approved / total)
        conversion_rate = (approved_quotes / total_quotes * 100) if total_quotes > 0 else 0
        
        return {
            "summary": {
                "total_quotes": total_quotes,
                "approved_quotes": approved_quotes,
                "pending_rfqs": pending_rfqs,
                "total_customers": total_customers,
                "new_customers_this_month": new_customers_this_month,
                "total_revenue": round(total_revenue, 2),
                "monthly_revenue": round(monthly_revenue, 2),
                "revenue_growth": round(revenue_growth, 1),
                "avg_quote_value": round(avg_quote_value, 2),
                "conversion_rate": round(conversion_rate, 1)
            }
        }
    except Exception as e:
        logging.error(f"Dashboard analytics error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch analytics: {str(e)}")


@router.get("/analytics/revenue-trend")
async def get_revenue_trend(
    months: int = 6,
    current_user: dict = Depends(get_current_user)
):
    """Get monthly revenue trend for the last N months"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        now = get_ist_now()
        trends = []
        
        for i in range(months - 1, -1, -1):
            # Calculate month start and end
            target_date = now - timedelta(days=30 * i)
            month_start = target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
            
            # Get revenue for this month
            pipeline = [
                {"$match": {
                    "status": "approved",
                    "created_at": {
                        "$gte": month_start.replace(tzinfo=None),
                        "$lt": month_end.replace(tzinfo=None)
                    }
                }},
                {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
            ]
            result = await db.quotes.aggregate(pipeline).to_list(1)
            revenue = result[0]["total"] if result else 0
            
            # Get quote count for this month
            quote_count = await db.quotes.count_documents({
                "created_at": {
                    "$gte": month_start.replace(tzinfo=None),
                    "$lt": month_end.replace(tzinfo=None)
                }
            })
            
            trends.append({
                "month": month_start.strftime("%b"),
                "year": month_start.year,
                "revenue": round(revenue, 2),
                "quotes": quote_count
            })
        
        return {"trends": trends}
    except Exception as e:
        logging.error(f"Revenue trend error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch revenue trend: {str(e)}")


@router.get("/analytics/top-customers")
async def get_top_customers(
    limit: int = 5,
    current_user: dict = Depends(get_current_user)
):
    """Get top customers by revenue"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        pipeline = [
            {"$match": {"status": "approved"}},
            {"$group": {
                "_id": "$customer_id",
                "total_revenue": {"$sum": "$total_price"},
                "quote_count": {"$sum": 1},
                "customer_name": {"$first": "$customer_name"},
                "company": {"$first": "$company"}
            }},
            {"$sort": {"total_revenue": -1}},
            {"$limit": limit}
        ]
        
        results = await db.quotes.aggregate(pipeline).to_list(limit)
        
        customers = []
        for r in results:
            customers.append({
                "customer_id": r["_id"],
                "customer_name": r.get("customer_name", "Unknown"),
                "company": r.get("company", "N/A"),
                "total_revenue": round(r["total_revenue"], 2),
                "quote_count": r["quote_count"]
            })
        
        return {"top_customers": customers}
    except Exception as e:
        logging.error(f"Top customers error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch top customers: {str(e)}")


@router.get("/analytics/quote-status")
async def get_quote_status_distribution(current_user: dict = Depends(get_current_user)):
    """Get distribution of quotes by status"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        pipeline = [
            {"$group": {
                "_id": {"$ifNull": ["$status", "pending"]},
                "count": {"$sum": 1}
            }}
        ]
        
        results = await db.quotes.aggregate(pipeline).to_list(10)
        
        distribution = {}
        for r in results:
            status = r["_id"] if r["_id"] else "pending"
            distribution[status] = r["count"]
        
        # Ensure both statuses exist
        if "approved" not in distribution:
            distribution["approved"] = 0
        if "pending" not in distribution:
            distribution["pending"] = 0
            
        return {"distribution": distribution}
    except Exception as e:
        logging.error(f"Quote status error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch quote status: {str(e)}")


@router.get("/analytics/recent-activity")
async def get_recent_activity(
    limit: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Get recent quotes and customer activity"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Recent quotes
        recent_quotes = await db.quotes.find(
            {},
            {"_id": 0, "quote_number": 1, "customer_name": 1, "company": 1, 
             "total_price": 1, "status": 1, "created_at": 1}
        ).sort("created_at", -1).limit(limit).to_list(limit)
        
        # Format dates
        for quote in recent_quotes:
            if quote.get("created_at"):
                quote["created_at"] = quote["created_at"].isoformat()
        
        # Recent customers
        recent_customers = await db.users.find(
            {"role": "customer"},
            {"_id": 0, "name": 1, "email": 1, "company": 1, "created_at": 1}
        ).sort("created_at", -1).limit(5).to_list(5)
        
        # Format dates
        for customer in recent_customers:
            if customer.get("created_at"):
                customer["created_at"] = customer["created_at"].isoformat()
        
        return {
            "recent_quotes": recent_quotes,
            "recent_customers": recent_customers
        }
    except Exception as e:
        logging.error(f"Recent activity error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch recent activity: {str(e)}")


@router.get("/analytics/roller-type-distribution")
async def get_roller_type_distribution(current_user: dict = Depends(get_current_user)):
    """Get distribution of roller types in quotes"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        pipeline = [
            {"$unwind": "$products"},
            {"$group": {
                "_id": "$products.roller_type",
                "count": {"$sum": 1},
                "total_value": {"$sum": "$products.price"}
            }},
            {"$sort": {"count": -1}}
        ]
        
        results = await db.quotes.aggregate(pipeline).to_list(10)
        
        distribution = []
        for r in results:
            if r["_id"]:
                distribution.append({
                    "roller_type": r["_id"],
                    "count": r["count"],
                    "total_value": round(r.get("total_value", 0), 2)
                })
        
        return {"distribution": distribution}
    except Exception as e:
        logging.error(f"Roller type distribution error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch roller type distribution: {str(e)}")


@router.get("/analytics/export/excel")
async def export_analytics_excel(current_user: dict = Depends(get_current_user)):
    """Export dashboard analytics to Excel file"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Fetch all analytics data
        now = get_ist_now()
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Summary stats
        total_quotes = await db.quotes.count_documents({})
        approved_quotes = await db.quotes.count_documents({"status": "approved"})
        pending_rfqs = await db.quotes.count_documents({"status": {"$ne": "approved"}})
        total_customers = await db.users.count_documents({"role": "customer"})
        
        revenue_pipeline = [
            {"$match": {"status": "approved"}},
            {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
        ]
        revenue_result = await db.quotes.aggregate(revenue_pipeline).to_list(1)
        total_revenue = revenue_result[0]["total"] if revenue_result else 0
        
        # Top customers
        top_customers_pipeline = [
            {"$match": {"status": "approved"}},
            {"$group": {
                "_id": "$customer_id",
                "total_revenue": {"$sum": "$total_price"},
                "quote_count": {"$sum": 1},
                "customer_name": {"$first": "$customer_name"},
                "company": {"$first": "$company"}
            }},
            {"$sort": {"total_revenue": -1}},
            {"$limit": 10}
        ]
        top_customers = await db.quotes.aggregate(top_customers_pipeline).to_list(10)
        
        # Recent quotes
        recent_quotes = await db.quotes.find(
            {},
            {"_id": 0, "quote_number": 1, "customer_name": 1, "company": 1, 
             "total_price": 1, "status": 1, "created_at": 1}
        ).sort("created_at", -1).limit(20).to_list(20)
        
        # Roller type distribution
        roller_pipeline = [
            {"$unwind": "$products"},
            {"$group": {
                "_id": "$products.roller_type",
                "count": {"$sum": 1},
                "total_value": {"$sum": "$products.price"}
            }},
            {"$sort": {"count": -1}}
        ]
        roller_types = await db.quotes.aggregate(roller_pipeline).to_list(10)
        
        # Create Excel workbook
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="960018", end_color="960018", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["Dashboard Analytics Report", ""],
            ["Generated on", now.strftime("%d %b %Y, %I:%M %p IST")],
            ["", ""],
            ["Metric", "Value"],
            ["Total Revenue", f"₹{total_revenue:,.2f}"],
            ["Total Quotes", total_quotes],
            ["Approved Quotes", approved_quotes],
            ["Pending RFQs", pending_rfqs],
            ["Total Customers", total_customers],
            ["Conversion Rate", f"{(approved_quotes/total_quotes*100) if total_quotes > 0 else 0:.1f}%"],
            ["Average Quote Value", f"₹{(total_revenue/approved_quotes) if approved_quotes > 0 else 0:,.2f}"],
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=16, color="960018")
                elif row_idx == 4:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                cell.border = thin_border
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        
        # Sheet 2: Top Customers
        ws_customers = wb.create_sheet("Top Customers")
        customer_headers = ["Rank", "Customer Name", "Company", "Total Revenue", "Quote Count"]
        
        for col_idx, header in enumerate(customer_headers, 1):
            cell = ws_customers.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, customer in enumerate(top_customers, 2):
            ws_customers.cell(row=row_idx, column=1, value=row_idx-1).border = thin_border
            ws_customers.cell(row=row_idx, column=2, value=customer.get("customer_name", "Unknown")).border = thin_border
            ws_customers.cell(row=row_idx, column=3, value=customer.get("company", "N/A")).border = thin_border
            ws_customers.cell(row=row_idx, column=4, value=f"₹{customer['total_revenue']:,.2f}").border = thin_border
            ws_customers.cell(row=row_idx, column=5, value=customer["quote_count"]).border = thin_border
        
        for col_idx in range(1, 6):
            ws_customers.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Sheet 3: Recent Quotes
        ws_quotes = wb.create_sheet("Recent Quotes")
        quote_headers = ["Quote Number", "Customer", "Company", "Total Price", "Status", "Date"]
        
        for col_idx, header in enumerate(quote_headers, 1):
            cell = ws_quotes.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, quote in enumerate(recent_quotes, 2):
            ws_quotes.cell(row=row_idx, column=1, value=quote.get("quote_number", "")).border = thin_border
            ws_quotes.cell(row=row_idx, column=2, value=quote.get("customer_name", "")).border = thin_border
            ws_quotes.cell(row=row_idx, column=3, value=quote.get("company", "")).border = thin_border
            ws_quotes.cell(row=row_idx, column=4, value=f"₹{quote.get('total_price', 0):,.2f}").border = thin_border
            ws_quotes.cell(row=row_idx, column=5, value=quote.get("status", "pending").title()).border = thin_border
            created_at = quote.get("created_at")
            date_str = created_at.strftime("%d %b %Y") if created_at else ""
            ws_quotes.cell(row=row_idx, column=6, value=date_str).border = thin_border
        
        for col_idx in range(1, 7):
            ws_quotes.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Sheet 4: Roller Type Distribution
        ws_rollers = wb.create_sheet("Roller Types")
        roller_headers = ["Roller Type", "Count", "Total Value"]
        
        for col_idx, header in enumerate(roller_headers, 1):
            cell = ws_rollers.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, roller in enumerate(roller_types, 2):
            if roller["_id"]:
                ws_rollers.cell(row=row_idx, column=1, value=roller["_id"]).border = thin_border
                ws_rollers.cell(row=row_idx, column=2, value=roller["count"]).border = thin_border
                ws_rollers.cell(row=row_idx, column=3, value=f"₹{roller.get('total_value', 0):,.2f}").border = thin_border
        
        for col_idx in range(1, 4):
            ws_rollers.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"Dashboard_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logging.error(f"Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")


@router.get("/analytics/export/pdf")
async def export_analytics_pdf(current_user: dict = Depends(get_current_user)):
    """Export dashboard analytics to PDF file"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        from fpdf import FPDF
        
        # Fetch all analytics data
        now = get_ist_now()
        
        # Summary stats
        total_quotes = await db.quotes.count_documents({})
        approved_quotes = await db.quotes.count_documents({"status": "approved"})
        pending_rfqs = await db.quotes.count_documents({"status": {"$ne": "approved"}})
        total_customers = await db.users.count_documents({"role": "customer"})
        
        revenue_pipeline = [
            {"$match": {"status": "approved"}},
            {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
        ]
        revenue_result = await db.quotes.aggregate(revenue_pipeline).to_list(1)
        total_revenue = revenue_result[0]["total"] if revenue_result else 0
        
        # Top customers
        top_customers_pipeline = [
            {"$match": {"status": "approved"}},
            {"$group": {
                "_id": "$customer_id",
                "total_revenue": {"$sum": "$total_price"},
                "quote_count": {"$sum": 1},
                "customer_name": {"$first": "$customer_name"},
                "company": {"$first": "$company"}
            }},
            {"$sort": {"total_revenue": -1}},
            {"$limit": 5}
        ]
        top_customers = await db.quotes.aggregate(top_customers_pipeline).to_list(5)
        
        # Recent quotes
        recent_quotes = await db.quotes.find(
            {},
            {"_id": 0, "quote_number": 1, "customer_name": 1, "total_price": 1, "status": 1, "created_at": 1}
        ).sort("created_at", -1).limit(10).to_list(10)
        
        # Create PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Header with Convero branding
        pdf.set_fill_color(150, 0, 24)  # Carmine Red
        pdf.rect(0, 0, 210, 40, 'F')
        pdf.set_text_color(255, 255, 255)
        
        # Company name
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_xy(10, 8)
        pdf.cell(0, 8, 'CONVERO SOLUTIONS')
        
        # Report title
        pdf.set_font('Helvetica', 'B', 18)
        pdf.set_xy(10, 18)
        pdf.cell(0, 10, 'Dashboard Analytics Report', align='C')
        
        # Generated timestamp
        pdf.set_font('Helvetica', '', 9)
        pdf.set_xy(10, 30)
        pdf.cell(0, 8, f'Report Generated: {now.strftime("%d %b %Y at %I:%M:%S %p IST")}', align='C')
        
        # Reset text color
        pdf.set_text_color(0, 0, 0)
        pdf.set_y(45)
        
        # Summary Section
        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 10, 'Summary', fill=True, ln=True)
        pdf.ln(5)
        
        # Summary metrics in a grid
        pdf.set_font('Helvetica', '', 11)
        col_width = 95
        
        def format_currency(val):
            if val >= 10000000:
                return f"Rs. {val/10000000:.2f} Cr"
            elif val >= 100000:
                return f"Rs. {val/100000:.2f} L"
            return f"Rs. {val:,.2f}"
        
        metrics = [
            ("Total Revenue", format_currency(total_revenue)),
            ("Total Quotes", str(total_quotes)),
            ("Approved Quotes", str(approved_quotes)),
            ("Pending RFQs", str(pending_rfqs)),
            ("Total Customers", str(total_customers)),
            ("Conversion Rate", f"{(approved_quotes/total_quotes*100) if total_quotes > 0 else 0:.1f}%"),
        ]
        
        for i, (label, value) in enumerate(metrics):
            if i % 2 == 0:
                x = 10
            else:
                x = 105
            
            pdf.set_x(x)
            pdf.set_font('Helvetica', '', 10)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(col_width, 6, label)
            
            if i % 2 == 1:
                pdf.ln()
            
            pdf.set_x(x)
            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(col_width, 8, value)
            
            if i % 2 == 1:
                pdf.ln(5)
        
        pdf.ln(10)
        
        # Top Customers Section
        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 10, 'Top Customers by Revenue', fill=True, ln=True)
        pdf.ln(3)
        
        # Table header
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_fill_color(150, 0, 24)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(10, 8, '#', border=1, fill=True, align='C')
        pdf.cell(60, 8, 'Customer', border=1, fill=True, align='C')
        pdf.cell(60, 8, 'Company', border=1, fill=True, align='C')
        pdf.cell(40, 8, 'Revenue', border=1, fill=True, align='C')
        pdf.cell(20, 8, 'Quotes', border=1, fill=True, align='C')
        pdf.ln()
        
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 9)
        for idx, customer in enumerate(top_customers, 1):
            pdf.cell(10, 7, str(idx), border=1, align='C')
            pdf.cell(60, 7, customer.get("customer_name", "Unknown")[:25], border=1)
            pdf.cell(60, 7, (customer.get("company", "N/A") or "N/A")[:25], border=1)
            pdf.cell(40, 7, f"Rs. {customer['total_revenue']:,.0f}", border=1, align='R')
            pdf.cell(20, 7, str(customer["quote_count"]), border=1, align='C')
            pdf.ln()
        
        pdf.ln(10)
        
        # Recent Quotes Section
        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 10, 'Recent Quotes', fill=True, ln=True)
        pdf.ln(3)
        
        # Table header
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_fill_color(150, 0, 24)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(40, 8, 'Quote #', border=1, fill=True, align='C')
        pdf.cell(50, 8, 'Customer', border=1, fill=True, align='C')
        pdf.cell(40, 8, 'Amount', border=1, fill=True, align='C')
        pdf.cell(30, 8, 'Status', border=1, fill=True, align='C')
        pdf.cell(30, 8, 'Date', border=1, fill=True, align='C')
        pdf.ln()
        
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 9)
        for quote in recent_quotes:
            pdf.cell(40, 7, quote.get("quote_number", "")[:15], border=1)
            pdf.cell(50, 7, (quote.get("customer_name", "")[:20]), border=1)
            pdf.cell(40, 7, f"Rs. {quote.get('total_price', 0):,.0f}", border=1, align='R')
            status = quote.get("status", "pending").title()
            pdf.cell(30, 7, status, border=1, align='C')
            created_at = quote.get("created_at")
            date_str = created_at.strftime("%d %b") if created_at else ""
            pdf.cell(30, 7, date_str, border=1, align='C')
            pdf.ln()
        
        # Footer
        pdf.set_y(-20)
        pdf.set_font('Helvetica', 'I', 8)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 10, 'Convero Solutions - Roller Price Calculator', align='C')
        
        # Output PDF
        pdf_bytes = pdf.output()
        filename = f"Dashboard_Report_{now.strftime('%Y%m%d_%H%M')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logging.error(f"PDF export error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export PDF: {str(e)}")

@router.post("/admin/migrate-customer-codes")
async def migrate_customer_codes(current_user: dict = Depends(get_current_user)):
    """Migrate existing customers to have customer codes - Admin only"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can run migrations")
    
    updated_count = 0
    
    # Find all users with role=customer and no customer_code
    users_cursor = db.users.find({
        "role": "customer",
        "$or": [
            {"customer_code": {"$exists": False}},
            {"customer_code": None}
        ]
    }).sort("created_at", 1)  # Sort by creation date to maintain order
    
    async for user in users_cursor:
        customer_code = await generate_customer_code()
        
        # Update user
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {"customer_code": customer_code}}
        )
        
        # Also update corresponding customer record
        await db.customers.update_one(
            {"email": user["email"]},
            {"$set": {"customer_code": customer_code}}
        )
        
        updated_count += 1
        logging.info(f"Assigned customer code {customer_code} to user {user['email']}")
    
    # Also update customers collection entries that don't have codes
    customers_cursor = db.customers.find({
        "$or": [
            {"customer_code": {"$exists": False}},
            {"customer_code": None}
        ]
    }).sort("created_at", 1)
    
    async for customer in customers_cursor:
        # Check if this customer's email has a user with a code
        user = await db.users.find_one({"email": customer.get("email"), "customer_code": {"$exists": True}})
        
        if user and user.get("customer_code"):
            # Use the same code as the user
            await db.customers.update_one(
                {"_id": customer["_id"]},
                {"$set": {"customer_code": user["customer_code"]}}
            )
        else:
            # Generate a new code
            customer_code = await generate_customer_code()
            await db.customers.update_one(
                {"_id": customer["_id"]},
                {"$set": {"customer_code": customer_code}}
            )
            updated_count += 1
            logging.info(f"Assigned customer code {customer_code} to customer {customer.get('email', customer.get('name'))}")
    
    # Also update quotes that don't have customer_code
    quotes_updated = 0
    quotes_cursor = db.quotes.find({
        "$or": [
            {"customer_code": {"$exists": False}},
            {"customer_code": None}
        ]
    })
    
    async for quote in quotes_cursor:
        # Try to find customer_code from customer_id or customer_email
        customer_code = None
        
        # First check by customer_id in users
        if quote.get("customer_id"):
            try:
                from bson import ObjectId
                user = await db.users.find_one({"_id": ObjectId(quote["customer_id"])})
                if user:
                    customer_code = user.get("customer_code")
            except:
                pass
        
        # If not found, try by email
        if not customer_code and quote.get("customer_email"):
            user = await db.users.find_one({"email": quote["customer_email"]})
            if user:
                customer_code = user.get("customer_code")
        
        # If still not found, check customers collection
        if not customer_code and quote.get("customer_email"):
            customer = await db.customers.find_one({"email": quote["customer_email"]})
            if customer:
                customer_code = customer.get("customer_code")
        
        if customer_code:
            await db.quotes.update_one(
                {"_id": quote["_id"]},
                {"$set": {"customer_code": customer_code}}
            )
            quotes_updated += 1
    
    return {"message": f"Migration complete. Updated {updated_count} customers and {quotes_updated} quotes with codes."}


