"""Inventory/Store Routes — Stock management, Purchase Orders, QC, Issue, Alerts"""
from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from bson import ObjectId
from routes import db, get_current_user, require_role, get_ist_now, get_financial_year, UserRole, format_date_dmy
from fastapi.responses import StreamingResponse
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import roller_standards as rs

router = APIRouter(prefix="/store")

PIPE_NOS_PER_METER = 1 / 6  # 1 nos = 6 m; conversely 1 m = 1/6 nos

def _pipe_weight_per_meter(stock_item: dict) -> Optional[float]:
    """Given a pipe stock item, return kg/meter (or None). Looks up
    PIPE_WEIGHT_PER_METER by diameter (from bom_match_key) and class (from name)."""
    if (stock_item or {}).get("category") != "pipe":
        return None
    name = stock_item.get("name", "") or ""
    cls = "B"
    if "(Class A)" in name: cls = "A"
    elif "(Class C)" in name: cls = "C"
    key = stock_item.get("bom_match_key", "") or ""
    try:
        dia = float(key.split(":")[1])
    except Exception:
        return None
    return rs.PIPE_WEIGHT_PER_METER.get(dia, {}).get(cls)


def _shaft_weight_per_meter(stock_item: dict) -> Optional[float]:
    if (stock_item or {}).get("category") != "shaft":
        return None
    key = stock_item.get("bom_match_key", "") or ""
    try:
        dia = int(float(key.split(":")[1]))
    except Exception:
        return None
    return rs.SHAFT_WEIGHT_PER_METER.get(dia)

STOCK_CATEGORIES = ["pipe", "shaft", "bearing", "housing", "seal", "circlip", "end_plate", "hub", "rubber_ring", "rubber_lagging", "grease", "paint", "other"]
PO_STATUSES = ["draft", "ordered", "partial_received", "received", "cancelled"]
QC_STATUSES = ["pending", "passed", "failed", "partial"]


# ============= MODELS =============

class StockItemCreate(BaseModel):
    name: str  # e.g., "Pipe 114.3mm x 4.5mm"
    category: str  # pipe, shaft, bearing, etc.
    unit_purchase: str = "meters"  # meters, kg, nos, sqm, litres
    unit_bom: str = "kg"  # kg, nos, sqm
    conversion_factor: float = 1.0
    current_stock: float = 0
    reorder_level: float = 0
    specifications: Optional[Dict[str, Any]] = None  # dia, thickness, material, bearing_no etc.
    # BOM match key — deterministic link to BOM components
    # Format: "{category}:{spec1}:{spec2}" e.g. "pipe:114.3:4.5", "shaft:25:EN-8", "bearing:6205", "housing:108/52"
    bom_match_key: Optional[str] = None


class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    items: List[Dict[str, Any]]  # [{stock_item_id, qty, rate, unit?, gst_rate?}]
    notes: Optional[str] = None
    expected_delivery: Optional[str] = None
    interstate: Optional[bool] = False   # IGST vs CGST+SGST split
    linked_wo_ids: Optional[List[str]] = None   # optional trace-back to the WO(s) this PO was raised for


class QCEntry(BaseModel):
    po_id: str
    item_index: int
    status: str  # passed, failed, partial
    accepted_qty: float = 0
    rejected_qty: float = 0
    reason: Optional[str] = None


class StockIssue(BaseModel):
    wo_id: str
    items: List[Dict[str, Any]]  # [{stock_item_id, qty, notes}]


# ============= HELPERS =============

async def generate_po_number():
    fy = get_financial_year()
    from routes import _next_seq, _max_suffix, format_number
    seed = await _max_suffix(db.purchase_orders, "po_number", f"PO/{fy}/")
    n = await _next_seq(f"po:{fy}", seed_value=seed)
    return await format_number("po", n)


# ============= STOCK ITEMS =============

@router.get("/items")
async def get_stock_items(category: Optional[str] = None, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    query = {}
    if category:
        query["category"] = category
    items = await db.stock_items.find(query, {"_id": 0}).sort("category", 1).to_list(1000)
    # Enrich with last purchase rate (from the most recent PO that includes this stock item)
    last_rate_map: dict = {}
    cursor = db.purchase_orders.find(
        {},
        {"_id": 0, "po_number": 1, "created_at": 1, "items": 1},
    ).sort("created_at", -1)
    async for po in cursor:
        created_at = po.get("created_at")
        po_number = po.get("po_number")
        for line in po.get("items") or []:
            sid = line.get("stock_item_id")
            if sid and sid not in last_rate_map and line.get("rate"):
                last_rate_map[sid] = {
                    "rate": line.get("rate"),
                    "po_number": po_number,
                    "date": created_at,
                    "unit": line.get("unit") or "",
                }
    for it in items:
        lr = last_rate_map.get(it.get("id"))
        it["last_purchase_rate"] = lr.get("rate") if lr else None
        it["last_purchase_po"] = lr.get("po_number") if lr else None
        it["last_purchase_date"] = lr.get("date") if lr else None
        it["last_purchase_unit"] = lr.get("unit") if lr else None
        # Pipe & Shaft: purchased by weight (kg). Pipe stock shown as nos (1 nos = 6 m).
        if it.get("category") == "pipe":
            wpm = _pipe_weight_per_meter(it)
            it["weight_per_meter_kg"] = wpm
            it["unit_purchase"] = "kg"
            it["stock_unit"] = "nos"
            cur_m = float(it.get("current_stock") or 0)
            it["current_stock_m"] = round(cur_m, 3)
            it["current_stock_nos"] = round(cur_m / 6, 3)
        elif it.get("category") == "shaft":
            wpm = _shaft_weight_per_meter(it)
            it["weight_per_meter_kg"] = wpm
            it["unit_purchase"] = "kg"
            it["stock_unit"] = "meters"
            cur_m = float(it.get("current_stock") or 0)
            it["current_stock_m"] = round(cur_m, 3)
    return {"items": items, "total": len(items)}


@router.post("/items")
async def create_stock_item(item: StockItemCreate, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    now = get_ist_now()
    doc = item.dict()
    doc.update({"id": str(ObjectId()), "created_by": current_user.get("email"), "created_at": now.isoformat()})
    await db.stock_items.insert_one(doc)
    del doc["_id"]
    return {"message": "Stock item created", "item": doc}


@router.put("/items/{item_id}")
async def update_stock_item(item_id: str, item: StockItemCreate, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    result = await db.stock_items.update_one({"id": item_id}, {"$set": {**item.dict(), "updated_at": get_ist_now().isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Stock item not found")
    return {"message": "Stock item updated"}


@router.delete("/items/{item_id}")
async def delete_stock_item(item_id: str, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    result = await db.stock_items.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stock item not found")
    return {"message": "Stock item deleted"}


# ============= PURCHASE ORDERS =============

@router.get("/purchase-orders")
async def get_purchase_orders(status: Optional[str] = None, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    query = {}
    if status:
        query["status"] = status
    pos = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    # Enrich with supplier name
    for po in pos:
        supplier = await db.suppliers_master.find_one({"id": po.get("supplier_id")}, {"_id": 0, "name": 1})
        po["supplier_name"] = supplier.get("name") if supplier else "Unknown"
    return {"purchase_orders": pos, "total": len(pos)}


@router.post("/purchase-orders")
async def create_purchase_order(po: PurchaseOrderCreate, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    now = get_ist_now()
    po_number = await generate_po_number()

    # Enrich items with stock item names + GST breakdown
    items = []
    subtotal = 0.0
    gst_total = 0.0
    cgst_total = 0.0
    sgst_total = 0.0
    igst_total = 0.0
    interstate = bool(po.interstate)
    for item in po.items:
        stock_item = await db.stock_items.find_one({"id": item.get("stock_item_id")}, {"_id": 0})
        qty = float(item.get("qty", 0) or 0)
        rate = float(item.get("rate", 0) or 0)
        gst_rate = float(item.get("gst_rate", 18) or 0)   # default 18% (standard capital-goods bracket)
        amount = round(qty * rate, 2)
        gst_amount = round(amount * gst_rate / 100, 2)
        cgst = 0.0
        sgst = 0.0
        igst = 0.0
        if interstate:
            igst = gst_amount
        else:
            cgst = round(gst_amount / 2, 2)
            sgst = round(gst_amount - cgst, 2)
        total_line = round(amount + gst_amount, 2)
        subtotal += amount
        gst_total += gst_amount
        cgst_total += cgst
        sgst_total += sgst
        igst_total += igst
        # Pipe & Shaft: purchase unit is always kg (auto-converts on receipt)
        line_unit = item.get("unit") or (stock_item.get("unit_purchase") if stock_item else "nos")
        if stock_item and stock_item.get("category") in ("pipe", "shaft"):
            line_unit = "kg"
        items.append({
            "stock_item_id": item.get("stock_item_id"),
            "stock_item_name": stock_item.get("name") if stock_item else "Unknown",
            "category": stock_item.get("category") if stock_item else "",
            "qty_ordered": qty,
            "rate": rate,
            "unit": line_unit,
            "amount": amount,
            "gst_rate": gst_rate,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "gst_amount": gst_amount,
            "total_line": total_line,
            "qty_received": 0,
            "qty_accepted": 0,
            "qty_rejected": 0,
            "qc_status": "pending",
        })

    # Resolve linked WOs → numbers (nice for display)
    linked_wo_numbers: List[str] = []
    if po.linked_wo_ids:
        async for w in db.work_orders.find({"id": {"$in": po.linked_wo_ids}}, {"_id": 0, "wo_number": 1}):
            if w.get("wo_number"):
                linked_wo_numbers.append(w["wo_number"])

    doc = {
        "id": str(ObjectId()),
        "po_number": po_number,
        "supplier_id": po.supplier_id,
        "items": items,
        "subtotal": round(subtotal, 2),
        "cgst_total": round(cgst_total, 2),
        "sgst_total": round(sgst_total, 2),
        "igst_total": round(igst_total, 2),
        "gst_total": round(gst_total, 2),
        "interstate": interstate,
        "total_amount": round(subtotal + gst_total, 2),
        "status": "ordered",
        "notes": po.notes,
        "expected_delivery": po.expected_delivery,
        "linked_wo_ids": po.linked_wo_ids or [],
        "linked_wo_numbers": linked_wo_numbers,
        "created_by": current_user.get("email"),
        "created_at": now.isoformat(),
    }
    await db.purchase_orders.insert_one(doc)
    del doc["_id"]
    return {"message": f"Purchase Order {po_number} created", "purchase_order": doc}


# ============= QC (Quality Check) =============

@router.post("/qc")
async def process_qc(qc: QCEntry, current_user: dict = Depends(require_role([UserRole.ADMIN, UserRole.QUALITY_INSPECTOR]))):
    po = await db.purchase_orders.find_one({"id": qc.po_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    items = po.get("items", [])
    if qc.item_index < 0 or qc.item_index >= len(items):
        raise HTTPException(status_code=400, detail="Invalid item index")

    now = get_ist_now()
    item = items[qc.item_index]
    item["qc_status"] = qc.status
    item["qty_received"] = qc.accepted_qty + qc.rejected_qty
    item["qty_accepted"] = qc.accepted_qty
    item["qty_rejected"] = qc.rejected_qty
    item["qc_reason"] = qc.reason
    item["qc_by"] = current_user.get("email")
    item["qc_at"] = now.isoformat()

    # If QC passed/partial, add accepted qty to stock
    if qc.accepted_qty > 0:
        stock_item = await db.stock_items.find_one({"id": item.get("stock_item_id")})
        if stock_item:
            # Pipe / Shaft purchased in kg — convert to meters (internal stock unit) before adding
            accepted_stock_qty = qc.accepted_qty
            receipt_note = ""
            if stock_item.get("category") == "pipe":
                wpm = _pipe_weight_per_meter(stock_item)
                if wpm and wpm > 0:
                    accepted_stock_qty = round(qc.accepted_qty / wpm, 3)
                    receipt_note = f" · {qc.accepted_qty} kg → {accepted_stock_qty} m ({round(accepted_stock_qty/6, 3)} nos)"
            elif stock_item.get("category") == "shaft":
                wpm = _shaft_weight_per_meter(stock_item)
                if wpm and wpm > 0:
                    accepted_stock_qty = round(qc.accepted_qty / wpm, 3)
                    receipt_note = f" · {qc.accepted_qty} kg → {accepted_stock_qty} m"
            new_stock = stock_item.get("current_stock", 0) + accepted_stock_qty
            await db.stock_items.update_one({"id": item["stock_item_id"]}, {"$set": {"current_stock": round(new_stock, 3)}})

            # Log transaction with WO attribution from the originating PO
            await db.stock_transactions.insert_one({
                "id": str(ObjectId()),
                "stock_item_id": item["stock_item_id"],
                "stock_item_name": item.get("stock_item_name"),
                "type": "in",
                "qty": accepted_stock_qty,
                "qty_purchased": qc.accepted_qty,
                "qty_purchased_unit": item.get("unit") or stock_item.get("unit_purchase") or "",
                "reference": f"PO: {po.get('po_number')} (QC {qc.status}){receipt_note}",
                "po_id": qc.po_id,
                "po_number": po.get("po_number"),
                "linked_wo_ids": po.get("linked_wo_ids") or [],
                "linked_wo_numbers": po.get("linked_wo_numbers") or [],
                "by": current_user.get("email"),
                "at": now.isoformat(),
            })

            # Pro-rata distribute receipt across linked WOs (if any) into a
            # lightweight `po_wo_receipts` ledger. This lets WorkOrder and the
            # Material-Status chip reflect that stock is now reserved for that WO.
            linked = po.get("linked_wo_ids") or []
            if linked:
                # Equal split across WOs (caller can manually issue as needed).
                share = round(accepted_stock_qty / len(linked), 3)
                for wo_id in linked:
                    await db.po_wo_receipts.insert_one({
                        "id": str(ObjectId()),
                        "po_id": qc.po_id,
                        "po_number": po.get("po_number"),
                        "wo_id": wo_id,
                        "stock_item_id": item["stock_item_id"],
                        "stock_item_name": item.get("stock_item_name"),
                        "qty_received": share,
                        "at": now.isoformat(),
                        "by": current_user.get("email"),
                    })

    # Update PO status
    all_qc = all(i.get("qc_status") in ["passed", "failed"] for i in items)
    any_received = any(i.get("qty_accepted", 0) > 0 for i in items)
    if all_qc:
        po_status = "received"
    elif any_received:
        po_status = "partial_received"
    else:
        po_status = "ordered"

    await db.purchase_orders.update_one({"_id": po["_id"]}, {"$set": {"items": items, "status": po_status, "updated_at": now.isoformat()}})

    return {"message": f"QC processed: {qc.status} ({qc.accepted_qty} accepted, {qc.rejected_qty} rejected)"}


# ============= STOCK ISSUE (against Work Order) =============

@router.post("/issue")
async def issue_stock(issue: StockIssue, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    wo = await db.work_orders.find_one({"id": issue.wo_id}, {"_id": 0, "wo_number": 1})
    if not wo:
        raise HTTPException(status_code=404, detail="Work Order not found")

    now = get_ist_now()
    issued_items = []
    insufficient = []

    for item in issue.items:
        stock_item = await db.stock_items.find_one({"id": item.get("stock_item_id")})
        if not stock_item:
            continue

        qty = item.get("qty", 0)
        current = stock_item.get("current_stock", 0)

        if current < qty:
            insufficient.append(f"{stock_item.get('name')}: need {qty}, have {current}")
            continue

        # Deduct stock
        new_stock = current - qty
        await db.stock_items.update_one({"id": item["stock_item_id"]}, {"$set": {"current_stock": round(new_stock, 3)}})

        # Log transaction
        await db.stock_transactions.insert_one({
            "id": str(ObjectId()),
            "stock_item_id": item["stock_item_id"],
            "stock_item_name": stock_item.get("name"),
            "type": "out",
            "qty": qty,
            "reference": f"WO: {wo.get('wo_number')}",
            "wo_id": issue.wo_id,
            "notes": item.get("notes", ""),
            "by": current_user.get("email"),
            "at": now.isoformat(),
        })

        issued_items.append({"name": stock_item.get("name"), "qty": qty})

    if insufficient:
        raise HTTPException(status_code=400, detail=f"Insufficient stock: {'; '.join(insufficient)}")

    return {"message": f"{len(issued_items)} items issued against {wo.get('wo_number')}", "issued": issued_items}


# ============= STOCK TRANSACTIONS LOG =============

@router.get("/transactions")
async def get_transactions(stock_item_id: Optional[str] = None, type: Optional[str] = None, limit: int = 100, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    query = {}
    if stock_item_id:
        query["stock_item_id"] = stock_item_id
    if type:
        query["type"] = type
    txns = await db.stock_transactions.find(query, {"_id": 0}).sort("at", -1).to_list(limit)
    return {"transactions": txns, "total": len(txns)}


# ============= LOW STOCK ALERTS =============

@router.get("/alerts")
async def get_low_stock_alerts(current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    # Find items where current_stock <= reorder_level
    items = await db.stock_items.find({"reorder_level": {"$gt": 0}}, {"_id": 0}).to_list(1000)
    alerts = []
    for item in items:
        if item.get("current_stock", 0) <= item.get("reorder_level", 0):
            alerts.append({
                "id": item.get("id"),
                "name": item.get("name"),
                "category": item.get("category"),
                "current_stock": item.get("current_stock", 0),
                "reorder_level": item.get("reorder_level", 0),
                "deficit": round(item.get("reorder_level", 0) - item.get("current_stock", 0), 3),
                "unit": item.get("unit_purchase"),
            })
    return {"alerts": alerts, "total": len(alerts)}


# ============= STORE DASHBOARD =============

@router.get("/dashboard")
async def get_store_dashboard(current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    total_items = await db.stock_items.count_documents({})
    total_pos = await db.purchase_orders.count_documents({})
    pending_pos = await db.purchase_orders.count_documents({"status": {"$in": ["ordered", "partial_received"]}})

    # Low stock count
    items = await db.stock_items.find({"reorder_level": {"$gt": 0}}, {"_id": 0, "current_stock": 1, "reorder_level": 1}).to_list(1000)
    low_stock_count = sum(1 for i in items if i.get("current_stock", 0) <= i.get("reorder_level", 0))

    # Category counts
    pipeline = [{"$group": {"_id": "$category", "count": {"$sum": 1}, "total_value": {"$sum": "$current_stock"}}}]
    categories = await db.stock_items.aggregate(pipeline).to_list(20)

    # Recent transactions
    recent = await db.stock_transactions.find({}, {"_id": 0}).sort("at", -1).to_list(10)

    return {
        "total_items": total_items,
        "total_pos": total_pos,
        "pending_pos": pending_pos,
        "low_stock_alerts": low_stock_count,
        "categories": {c["_id"]: c["count"] for c in categories if c["_id"]},
        "recent_transactions": recent,
    }


# ============= OPTIONS =============

@router.get("/options")
async def get_store_options(current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    return {
        "categories": STOCK_CATEGORIES,
        "po_statuses": PO_STATUSES,
        "qc_statuses": QC_STATUSES,
        "units": ["meters", "kg", "nos", "sqm", "litres", "sets"],
    }


# ============= EXCEL EXPORTS =============

def _make_header(ws, headers, row=1):
    hf = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18


@router.get("/export/stock")
async def export_stock_excel(current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    items = await db.stock_items.find({}, {"_id": 0}).sort("category", 1).to_list(1000)
    wb = Workbook(); ws = wb.active; ws.title = "Stock"
    _make_header(ws, ["Name", "Category", "Current Stock", "Unit (Purchase)", "Unit (BOM)", "Reorder Level", "Status"])
    for r, item in enumerate(items, 2):
        ws.cell(row=r, column=1, value=item.get("name"))
        ws.cell(row=r, column=2, value=item.get("category"))
        ws.cell(row=r, column=3, value=item.get("current_stock", 0))
        ws.cell(row=r, column=4, value=item.get("unit_purchase"))
        ws.cell(row=r, column=5, value=item.get("unit_bom"))
        ws.cell(row=r, column=6, value=item.get("reorder_level", 0))
        low = item.get("current_stock", 0) <= item.get("reorder_level", 0) and item.get("reorder_level", 0) > 0
        ws.cell(row=r, column=7, value="LOW STOCK" if low else "OK")
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": "attachment; filename=Stock_Inventory.xlsx"})


@router.get("/export/purchase-orders")
async def export_po_excel(current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    pos = await db.purchase_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    wb = Workbook(); ws = wb.active; ws.title = "Purchase Orders"
    _make_header(ws, ["PO Number", "Supplier", "Item", "Qty Ordered", "Rate", "Amount", "Qty Accepted", "Qty Rejected", "QC Status", "PO Status", "Date"])
    r = 2
    for po in pos:
        supplier = await db.suppliers_master.find_one({"id": po.get("supplier_id")}, {"_id": 0, "name": 1})
        sup_name = supplier.get("name") if supplier else "Unknown"
        for item in po.get("items", []):
            ws.cell(row=r, column=1, value=po.get("po_number"))
            ws.cell(row=r, column=2, value=sup_name)
            ws.cell(row=r, column=3, value=item.get("stock_item_name"))
            ws.cell(row=r, column=4, value=item.get("qty_ordered"))
            ws.cell(row=r, column=5, value=item.get("rate"))
            ws.cell(row=r, column=6, value=item.get("amount"))
            ws.cell(row=r, column=7, value=item.get("qty_accepted", 0))
            ws.cell(row=r, column=8, value=item.get("qty_rejected", 0))
            ws.cell(row=r, column=9, value=item.get("qc_status"))
            ws.cell(row=r, column=10, value=po.get("status"))
            ws.cell(row=r, column=11, value=format_date_dmy(po.get("created_at")))
            r += 1
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": "attachment; filename=Purchase_Orders.xlsx"})


@router.get("/export/transactions")
async def export_transactions_excel(current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    txns = await db.stock_transactions.find({}, {"_id": 0}).sort("at", -1).to_list(5000)
    wb = Workbook(); ws = wb.active; ws.title = "Transactions"
    _make_header(ws, ["Item", "Type", "Qty", "Reference", "Notes", "By", "Date"])
    for r, t in enumerate(txns, 2):
        ws.cell(row=r, column=1, value=t.get("stock_item_name"))
        ws.cell(row=r, column=2, value=t.get("type", "").upper())
        ws.cell(row=r, column=3, value=t.get("qty"))
        ws.cell(row=r, column=4, value=t.get("reference"))
        ws.cell(row=r, column=5, value=t.get("notes", ""))
        ws.cell(row=r, column=6, value=t.get("by"))
        ws.cell(row=r, column=7, value=format_date_dmy(t.get("at")))
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": "attachment; filename=Stock_Transactions.xlsx"})


@router.get("/export/suppliers")
async def export_suppliers_excel(current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    suppliers = await db.suppliers_master.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    wb = Workbook(); ws = wb.active; ws.title = "Suppliers"
    _make_header(ws, ["Name", "Contact Person", "Phone", "Email", "GST", "City", "State", "Payment Terms"])
    for r, s in enumerate(suppliers, 2):
        ws.cell(row=r, column=1, value=s.get("name"))
        ws.cell(row=r, column=2, value=s.get("contact_person"))
        ws.cell(row=r, column=3, value=s.get("phone"))
        ws.cell(row=r, column=4, value=s.get("email"))
        ws.cell(row=r, column=5, value=s.get("gst_number"))
        ws.cell(row=r, column=6, value=s.get("city"))
        ws.cell(row=r, column=7, value=s.get("state"))
        ws.cell(row=r, column=8, value=s.get("payment_terms"))
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": "attachment; filename=Suppliers.xlsx"})


# ============= STOCK ADJUSTMENT =============

class StockAdjustment(BaseModel):
    stock_item_id: str
    adjustment_qty: float  # positive = add, negative = deduct
    reason: str  # damage, audit, opening_balance, correction, other
    notes: Optional[str] = None


@router.post("/adjust")
async def adjust_stock(adj: StockAdjustment, current_user: dict = Depends(require_role([UserRole.ADMIN]))):
    stock_item = await db.stock_items.find_one({"id": adj.stock_item_id})
    if not stock_item:
        raise HTTPException(status_code=404, detail="Stock item not found")

    now = get_ist_now()
    old_stock = stock_item.get("current_stock", 0)
    new_stock = round(old_stock + adj.adjustment_qty, 3)
    if new_stock < 0:
        raise HTTPException(status_code=400, detail=f"Stock cannot go below 0. Current: {old_stock}, Adjustment: {adj.adjustment_qty}")

    await db.stock_items.update_one({"id": adj.stock_item_id}, {"$set": {"current_stock": new_stock}})

    # Log transaction
    await db.stock_transactions.insert_one({
        "id": str(ObjectId()),
        "stock_item_id": adj.stock_item_id,
        "stock_item_name": stock_item.get("name"),
        "type": "adjust_in" if adj.adjustment_qty > 0 else "adjust_out",
        "qty": abs(adj.adjustment_qty),
        "reference": f"Adjustment: {adj.reason}",
        "notes": adj.notes or "",
        "by": current_user.get("email"),
        "at": now.isoformat(),
    })

    return {"message": f"Stock adjusted: {stock_item.get('name')} — {old_stock} → {new_stock}", "new_stock": new_stock}


# ============= PURCHASE ORDER PDF =============

@router.get("/purchase-orders/{po_id}/pdf")
async def get_po_pdf(po_id: str, token: Optional[str] = None, authorization: Optional[str] = None):
    from fastapi import Header
    from jose import jwt
    from routes import SECRET_KEY, ALGORITHM

    auth_token = token
    if not auth_token and authorization and authorization.startswith("Bearer "):
        auth_token = authorization[7:]
    if not auth_token:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        if not payload.get("sub"):
            raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    supplier = await db.suppliers_master.find_one({"id": po.get("supplier_id")}, {"_id": 0})
    sup_name = supplier.get("name", "") if supplier else ""
    sup_contact = supplier.get("contact_person", "") if supplier else ""
    sup_phone = supplier.get("phone", "") if supplier else ""
    sup_gst = supplier.get("gst_number", "") if supplier else ""
    sup_city = supplier.get("city", "") if supplier else ""
    sup_address = supplier.get("address", "") if supplier else ""

    import os
    COMPANY = {
        "name": os.environ.get("COMPANY_NAME", "CONVERO SOLUTIONS"),
        "address": os.environ.get("COMPANY_ADDRESS", ""),
        "phone": os.environ.get("COMPANY_PHONE", ""),
        "email": os.environ.get("COMPANY_EMAIL", ""),
        "gstin": os.environ.get("COMPANY_GSTIN", ""),
    }

    item_rows = ""
    total = 0
    for i, item in enumerate(po.get("items", []), 1):
        amt = item.get("amount", 0)
        total += amt
        item_rows += f"""<tr>
            <td style="text-align:center">{i}</td>
            <td>{item.get('stock_item_name','')}</td>
            <td style="text-align:center">{item.get('qty_ordered','')}</td>
            <td style="text-align:center">{item.get('unit','')}</td>
            <td style="text-align:right">Rs.{item.get('rate',0):,.2f}</td>
            <td style="text-align:right"><b>Rs.{amt:,.2f}</b></td>
        </tr>"""

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
    @page {{ size: A4; margin: 15mm; }}
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family: 'Helvetica Neue', Arial, sans-serif; color:#1E293B; font-size:11px; }}
    .po {{ max-width:800px; margin:0 auto; }}
    .header {{ display:flex; justify-content:space-between; border-bottom:3px solid #C5964A; padding-bottom:14px; margin-bottom:16px; }}
    .company-name {{ font-size:22px; font-weight:800; color:#0F172A; }}
    .company-details {{ font-size:10px; color:#475569; line-height:1.6; }}
    .doc-title {{ font-size:22px; font-weight:800; color:#960018; text-align:right; }}
    .info-grid {{ display:flex; gap:16px; margin-bottom:16px; }}
    .info-box {{ flex:1; background:#F8FAFC; border:1px solid #E2E8F0; border-radius:6px; padding:12px; }}
    .info-label {{ font-size:9px; font-weight:700; color:#C5964A; letter-spacing:1px; text-transform:uppercase; margin-bottom:4px; }}
    .info-value {{ font-size:11px; color:#0F172A; }}
    table.items {{ width:100%; border-collapse:collapse; margin-bottom:16px; }}
    table.items th {{ background:#0F172A; color:#fff; padding:8px; font-size:10px; }}
    table.items td {{ padding:8px; border-bottom:1px solid #E2E8F0; }}
    .total-row {{ font-size:14px; font-weight:800; text-align:right; color:#960018; padding:10px 0; border-top:2px solid #C5964A; }}
    .terms {{ font-size:10px; color:#64748B; margin-top:16px; }}
    .stamp-area {{ display:flex; justify-content:space-between; margin-top:40px; }}
    .stamp-line {{ width:180px; border-top:1px solid #CBD5E1; margin-top:50px; padding-top:4px; font-size:10px; color:#64748B; text-align:center; }}
    .footer {{ text-align:center; margin-top:20px; padding-top:10px; border-top:1px solid #E2E8F0; font-size:9px; color:#94A3B8; }}
</style></head><body>
<div class="po">
    <div class="header">
        <div><div class="company-name">{COMPANY['name']}</div><div class="company-details">{COMPANY['address']}<br>Ph: {COMPANY['phone']} | {COMPANY['email']}<br><b>GSTIN: {COMPANY['gstin']}</b></div></div>
        <div><div class="doc-title">PURCHASE ORDER</div><div style="font-size:14px;font-weight:600;text-align:right">{po.get('po_number','')}</div><div style="font-size:11px;color:#64748B;text-align:right">Date: {format_date_dmy(po.get('created_at'))}</div></div>
    </div>
    <div class="info-grid">
        <div class="info-box"><div class="info-label">Supplier</div><div class="info-value"><b>{sup_name}</b><br>{sup_contact}<br>{sup_phone}<br>{sup_city} {sup_address}<br>{'GSTIN: ' + sup_gst if sup_gst else ''}</div></div>
        <div class="info-box"><div class="info-label">PO Details</div><div class="info-value">Status: {po.get('status','')}<br>{'Expected: ' + po.get('expected_delivery','') if po.get('expected_delivery') else ''}<br>{'Notes: ' + po.get('notes','') if po.get('notes') else ''}</div></div>
    </div>
    <table class="items"><tr><th>Sr.</th><th>Item</th><th>Qty</th><th>Unit</th><th style="text-align:right">Rate</th><th style="text-align:right">Amount</th></tr>{item_rows}</table>
    <div class="total-row">Total: Rs.{total:,.2f}</div>
    <div class="terms"><b>Terms:</b><br>1. Material must conform to IS standards.<br>2. Delivery as per schedule mentioned.<br>3. Payment as per agreed terms.<br>4. Subject to Ahmedabad jurisdiction.</div>
    <div class="stamp-area"><div><div class="stamp-line">Supplier's Acceptance</div></div><div><div class="stamp-line">For {COMPANY['name']}<br>Authorized Signatory</div></div></div>
    <div class="footer">{COMPANY['name']} | GSTIN: {COMPANY['gstin']} | {COMPANY['email']}</div>
</div></body></html>"""

    output = io.BytesIO(html.encode('utf-8'))
    output.seek(0)
    return StreamingResponse(output, media_type="text/html",
                           headers={"Content-Disposition": f"attachment; filename={po.get('po_number','PO').replace('/','_')}.html"})
