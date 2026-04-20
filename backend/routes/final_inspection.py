"""Final Inspection Report
Consolidates Pipe WIP QC + Shaft WIP QC + post-assembly tests (Runout, Water, Dust,
Friction, Painting/DFT, Bearing match, Rust preventive, Welding) into a single
per-WO sign-off sheet. One record per (WO, product index, sample). Sample count is
locked to the Pipe WIP QC sample_qty per product (falls back to Shaft QC).

Permissions: Admin, Production Head, Quality Inspector.
"""
from fastapi import APIRouter, HTTPException, Depends, Header, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import io
import os
from jose import jwt

from routes import (db, require_role, UserRole, SECRET_KEY, ALGORITHM, format_date_dmy)

router = APIRouter()


# ─────────────────────────────── Models ────────────────────────────────────

class FinalInspectionSample(BaseModel):
    sample_no: int
    runout_mm: Optional[float] = None
    water_ok: Optional[bool] = None            # Pass/Fail
    dust_ok: Optional[bool] = None             # Pass/Fail
    friction_coeff: Optional[float] = None     # < 0.02 = pass
    painting_visual_ok: Optional[bool] = None  # satisfactory / not
    dft_microns: Optional[float] = None
    bearing_match: Optional[bool] = None       # Y/N vs WO spec
    bearing_reason: Optional[str] = None
    rust_preventive: Optional[bool] = None     # Y/N
    rust_reason: Optional[str] = None
    welding_ok: Optional[bool] = None          # satisfactory / not
    remarks: Optional[str] = None


class FinalInspectionItemRecord(BaseModel):
    item_index: int
    expected_dft_microns: Optional[float] = None  # entered once per item
    samples: List[FinalInspectionSample]


class FinalInspectionRequest(BaseModel):
    items: List[FinalInspectionItemRecord]


# ───────────────────────────── Tolerance helpers ───────────────────────────

def _runout_tolerance(pipe_length_mm: float) -> float:
    """≤1350mm → 1.6, else 2.0 (user spec)."""
    if not pipe_length_mm:
        return 1.6
    return 1.6 if float(pipe_length_mm) <= 1350 else 2.0


def _dft_in_tolerance(measured: float, expected: float) -> bool:
    if not expected or measured is None:
        return False
    tol = abs(float(expected)) * 0.20
    return abs(float(measured) - float(expected)) <= tol


def _evaluate_sample(s: dict, expected_dft: Optional[float], pipe_length_mm: float, applicable: Optional[dict] = None) -> dict:
    """Return sample dict enriched with *_ok flags and overall_pass.
    Tests marked non-applicable via *applicable* dict are ignored in overall_pass.
    Applicable keys: runout, water, dust, friction, painting (default True each).
    Bearing match, Rust preventive, Welding are always applicable."""
    ap = applicable or {}
    app_runout = ap.get("runout", True) if ap else True
    app_water = ap.get("water", True) if ap else True
    app_dust = ap.get("dust", True) if ap else True
    app_friction = ap.get("friction", True) if ap else True
    app_painting = ap.get("painting", True) if ap else True

    out = dict(s)

    # Runout
    runout_tol = _runout_tolerance(pipe_length_mm)
    r = s.get("runout_mm")
    out["runout_tolerance_mm"] = runout_tol
    out["runout_ok"] = (r is not None) and (float(r) < runout_tol)

    # Friction
    f = s.get("friction_coeff")
    out["friction_ok"] = (f is not None) and (float(f) < 0.02)

    # DFT
    dft = s.get("dft_microns")
    out["dft_ok"] = _dft_in_tolerance(dft, expected_dft) if (dft is not None and expected_dft) else False

    # Yes/No fields already pass/fail bools
    def _b(k): return bool(s.get(k)) if s.get(k) is not None else False
    water_ok = _b("water_ok")
    dust_ok = _b("dust_ok")
    paint_ok = _b("painting_visual_ok")
    bearing_ok = _b("bearing_match")
    rust_ok = _b("rust_preventive")
    weld_ok = _b("welding_ok")

    # Overall = AND of ticked tests + always-on tests (bearing, rust, weld)
    checks = []
    if app_runout: checks.append(out["runout_ok"])
    if app_water: checks.append(water_ok)
    if app_dust: checks.append(dust_ok)
    if app_friction: checks.append(out["friction_ok"])
    if app_painting:
        checks.append(paint_ok)
        checks.append(out["dft_ok"])
    checks.append(bearing_ok)
    checks.append(rust_ok)
    checks.append(weld_ok)

    out["overall_pass"] = all(checks) if checks else False
    return out


# ───────────────────────────── Context helpers ─────────────────────────────

async def _build_wo_context(wo: dict) -> dict:
    """Derive per-item context: pipe length, roller length tolerance, bearing spec,
    sample quantity (from Pipe WIP QC, fall back to Shaft WIP QC). Also resolve
    the test_requirements from the parent SO (5 optional post-assembly tests)."""
    wo_id = wo.get("id")

    # Resolve applicable tests from parent SO (fall back: all 5 applicable)
    so = None
    if wo.get("so_id"):
        so = await db.sales_orders.find_one({"id": wo["so_id"]}, {"_id": 0, "test_requirements": 1})
    if not so and wo.get("so_number"):
        so = await db.sales_orders.find_one({"so_number": wo["so_number"]}, {"_id": 0, "test_requirements": 1})
    so_tests = (so or {}).get("test_requirements") or {}
    applicable = {
        "runout": bool(so_tests.get("runout", True)) if so_tests else True,
        "water": bool(so_tests.get("water", True)) if so_tests else True,
        "dust": bool(so_tests.get("dust", True)) if so_tests else True,
        "friction": bool(so_tests.get("friction_factor", True)) if so_tests else True,
        "painting": bool(so_tests.get("painting", True)) if so_tests else True,
    }

    sub_wos = await db.sub_work_orders.find({"parent_wo_id": wo_id}, {"_id": 0}).to_list(None)
    pipe_sub = next((s for s in sub_wos if s.get("type") == "pipe"), None)
    shaft_sub = next((s for s in sub_wos if s.get("type") == "shaft"), None)

    pipe_qty_by_idx: dict = {}
    shaft_qty_by_idx: dict = {}
    pipe_items_by_idx: dict = {}
    shaft_items_by_idx: dict = {}
    if pipe_sub:
        for rec in (pipe_sub.get("wip_qc") or {}).get("items", []) or []:
            idx = int(rec.get("item_index"))
            pipe_qty_by_idx[idx] = int(rec.get("sample_qty") or 0)
            pipe_items_by_idx[idx] = rec
    if shaft_sub:
        for rec in (shaft_sub.get("wip_qc") or {}).get("items", []) or []:
            idx = int(rec.get("item_index"))
            shaft_qty_by_idx[idx] = int(rec.get("sample_qty") or 0)
            shaft_items_by_idx[idx] = rec

    pipe_status = ((pipe_sub or {}).get("wip_qc") or {}).get("status") or "none"
    shaft_status = ((shaft_sub or {}).get("wip_qc") or {}).get("status") or "none"

    items_ctx = []
    wo_items = wo.get("items") or wo.get("products") or []
    for idx, p in enumerate(wo_items):
        specs = p.get("specifications") or {}
        pipe_length = specs.get("pipe_length") or 0
        if not pipe_length and pipe_sub:
            psub_items = pipe_sub.get("items") or []
            if idx < len(psub_items):
                pipe_length = psub_items[idx].get("pipe_length") or 0
        bearing_number = specs.get("bearing_number") or specs.get("bearing") or "-"
        bearing_make = specs.get("bearing_make") or "China"
        sample_qty = pipe_qty_by_idx.get(idx) or shaft_qty_by_idx.get(idx) or 0
        items_ctx.append({
            "item_index": idx,
            "product_name": p.get("product_name"),
            "product_code": p.get("product_code") or p.get("product_id"),
            "quantity": p.get("quantity", 1),
            "pipe_length_mm": float(pipe_length or 0),
            "runout_tolerance_mm": _runout_tolerance(float(pipe_length or 0)),
            "bearing_spec": f"{bearing_number} {bearing_make}".strip(),
            "sample_qty": int(sample_qty),
            "sample_source": "pipe" if idx in pipe_qty_by_idx else ("shaft" if idx in shaft_qty_by_idx else None),
            "pipe_wip_qc": pipe_items_by_idx.get(idx),
            "shaft_wip_qc": shaft_items_by_idx.get(idx),
        })
    return {
        "wo_number": wo.get("wo_number"),
        "so_number": wo.get("so_number"),
        "customer_name": wo.get("customer_name"),
        "customer_company": wo.get("customer_company"),
        "delivery_date": wo.get("delivery_date"),
        "applicable_tests": applicable,
        "pipe_qc_status": pipe_status,
        "shaft_qc_status": shaft_status,
        "items": items_ctx,
    }


# ───────────────────────────── Endpoints ───────────────────────────────────

@router.get("/work-orders/{wo_id}/final-inspection")
async def get_final_inspection(
    wo_id: str,
    current_user: dict = Depends(require_role([UserRole.ADMIN, UserRole.PRODUCTION_HEAD, UserRole.QUALITY_INSPECTOR])),
):
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work Order not found")
    ctx = await _build_wo_context(wo)
    existing = await db.final_inspection_records.find_one({"wo_id": wo_id}, {"_id": 0})
    return {"context": ctx, "record": existing}


@router.get("/final-inspection/gate-overview")
async def final_inspection_gate_overview(
    current_user: dict = Depends(require_role(UserRole.all_staff())),
):
    """Lightweight aggregation used by WO cards: for every WO, returns its
    applicable_tests (from parent SO), pipe/shaft WIP QC status, and FIP status.
    """
    wos = await db.work_orders.find({}, {"_id": 0, "id": 1, "so_id": 1, "so_number": 1}).to_list(2000)
    subs = await db.sub_work_orders.find({}, {"_id": 0}).to_list(5000)
    sub_by_parent: dict = {}
    for s in subs:
        sub_by_parent.setdefault(s.get("parent_wo_id"), []).append(s)
    fis = await db.final_inspection_records.find({}, {"_id": 0, "wo_id": 1, "status": 1}).to_list(3000)
    fi_by_wo = {f.get("wo_id"): f.get("status") for f in fis}

    # Build SO → test_requirements cache (only fetch once)
    so_ids = {w.get("so_id") for w in wos if w.get("so_id")}
    so_numbers = {w.get("so_number") for w in wos if not w.get("so_id") and w.get("so_number")}
    so_cache: dict = {}
    if so_ids:
        async for so in db.sales_orders.find({"id": {"$in": list(so_ids)}}, {"_id": 0, "id": 1, "test_requirements": 1}):
            so_cache[so.get("id")] = so.get("test_requirements") or {}
    if so_numbers:
        async for so in db.sales_orders.find({"so_number": {"$in": list(so_numbers)}}, {"_id": 0, "so_number": 1, "test_requirements": 1}):
            so_cache[so.get("so_number")] = so.get("test_requirements") or {}

    out = {}
    for w in wos:
        wo_id = w.get("id")
        tr = so_cache.get(w.get("so_id")) or so_cache.get(w.get("so_number")) or {}
        applicable = {
            "runout": bool(tr.get("runout", True)),
            "water": bool(tr.get("water", True)),
            "dust": bool(tr.get("dust", True)),
            "friction": bool(tr.get("friction_factor", True)),
            "painting": bool(tr.get("painting", True)),
        }
        wo_subs = sub_by_parent.get(wo_id) or []
        pipe_sub = next((s for s in wo_subs if s.get("type") == "pipe"), None)
        shaft_sub = next((s for s in wo_subs if s.get("type") == "shaft"), None)
        pipe_status = ((pipe_sub or {}).get("wip_qc") or {}).get("status") or "none"
        shaft_status = ((shaft_sub or {}).get("wip_qc") or {}).get("status") or "none"
        out[wo_id] = {
            "applicable_tests": applicable,
            "pipe_qc_status": pipe_status,
            "shaft_qc_status": shaft_status,
            "fi_status": fi_by_wo.get(wo_id) or "none",
            "fi_eligible": pipe_status == "passed" and shaft_status == "passed",
        }
    return {"by_wo_id": out}


@router.post("/work-orders/{wo_id}/final-inspection")
async def save_final_inspection(
    wo_id: str,
    data: dict,
    current_user: dict = Depends(require_role([UserRole.ADMIN, UserRole.PRODUCTION_HEAD, UserRole.QUALITY_INSPECTOR])),
):
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work Order not found")
    ctx = await _build_wo_context(wo)
    # Gate: both Pipe and Shaft WIP QC must be PASSED before final inspection
    if ctx.get("pipe_qc_status") != "passed" or ctx.get("shaft_qc_status") != "passed":
        raise HTTPException(
            status_code=400,
            detail=f"Final Inspection blocked — Pipe WIP QC: {ctx.get('pipe_qc_status')} · Shaft WIP QC: {ctx.get('shaft_qc_status')}. Both must be PASSED."
        )
    ctx_by_idx = {c["item_index"]: c for c in ctx["items"]}
    applicable = ctx.get("applicable_tests") or {}

    records_in = (data or {}).get("items") or []
    enriched_items = []
    any_fail = False

    for rec in records_in:
        idx = rec.get("item_index")
        if idx is None or idx not in ctx_by_idx:
            continue
        item_ctx = ctx_by_idx[idx]
        locked_qty = int(item_ctx.get("sample_qty") or 0)
        expected_dft = rec.get("expected_dft_microns")

        samples_out = []
        pass_count = 0
        for s in (rec.get("samples") or [])[:locked_qty] if locked_qty else (rec.get("samples") or []):
            sd = _evaluate_sample(s, expected_dft, item_ctx["pipe_length_mm"], applicable)
            samples_out.append(sd)
            if sd["overall_pass"]:
                pass_count += 1
            else:
                any_fail = True

        enriched_items.append({
            "item_index": idx,
            "product_name": item_ctx["product_name"],
            "product_code": item_ctx["product_code"],
            "quantity": item_ctx["quantity"],
            "pipe_length_mm": item_ctx["pipe_length_mm"],
            "runout_tolerance_mm": item_ctx["runout_tolerance_mm"],
            "bearing_spec": item_ctx["bearing_spec"],
            "sample_qty": locked_qty,
            "expected_dft_microns": expected_dft,
            "samples": samples_out,
            "pass_count": pass_count,
            "fail_count": len(samples_out) - pass_count,
        })

    now = datetime.now(timezone.utc).isoformat()
    status = "passed" if (enriched_items and not any_fail) else ("failed" if any_fail else "pending")
    doc = {
        "wo_id": wo_id,
        "wo_number": wo.get("wo_number"),
        "items": enriched_items,
        "applicable_tests": applicable,
        "status": status,
        "inspected_by": current_user.get("email"),
        "inspector_name": current_user.get("full_name") or current_user.get("email"),
        "inspector_role": current_user.get("role"),
        "inspected_at": now,
        "updated_at": now,
    }
    await db.final_inspection_records.update_one(
        {"wo_id": wo_id}, {"$set": doc, "$setOnInsert": {"created_at": now}}, upsert=True,
    )
    saved = await db.final_inspection_records.find_one({"wo_id": wo_id}, {"_id": 0})
    return {"message": f"Final Inspection saved — status: {status.upper()}", "record": saved}


# ───────────────────────────── PDF Report ──────────────────────────────────

COMPANY = {
    "name": os.environ.get("COMPANY_NAME", "CONVERO SOLUTIONS"),
    "address": os.environ.get("COMPANY_ADDRESS", ""),
    "phone": os.environ.get("COMPANY_PHONE", ""),
    "email": os.environ.get("COMPANY_EMAIL", ""),
}


def _yn(v) -> str:
    if v is True: return "Yes"
    if v is False: return "No"
    return "–"


def _pf(v) -> str:
    if v is True: return '<span style="color:#065F46;font-weight:700">PASS</span>'
    if v is False: return '<span style="color:#B91C1C;font-weight:700">FAIL</span>'
    return "–"


def _fmt(v, suffix=""):
    if v is None or v == "": return "–"
    try:
        return f"{float(v):.3g}{suffix}"
    except Exception:
        return str(v)


def _render_final_inspection_html(record: dict, wo: dict, ctx: Optional[dict] = None) -> str:
    status = (record.get("status") or "pending").upper()
    stamp_color = "#059669" if status == "PASSED" else ("#DC2626" if status == "FAILED" else "#6B7280")

    applicable = record.get("applicable_tests") or {"runout": True, "water": True, "dust": True, "friction": True, "painting": True}
    show_runout = bool(applicable.get("runout", True))
    show_water = bool(applicable.get("water", True))
    show_dust = bool(applicable.get("dust", True))
    show_friction = bool(applicable.get("friction", True))
    show_painting = bool(applicable.get("painting", True))

    # Build dynamic header row
    head_cols = ["#"]
    if show_runout: head_cols.append("Runout")
    if show_water: head_cols.append("Water")
    if show_dust: head_cols.append("Dust")
    if show_friction: head_cols.append("Friction<br>(COF &lt; 0.02)")
    if show_painting: head_cols += ["Paint<br>Visual", "DFT<br>(±20%)"]
    head_cols += ["Bearing<br>Match", "Rust<br>Prev.", "Welding", "Overall"]
    head_html = "".join(f"<th>{c}</th>" for c in head_cols)

    # Build lookup of pipe/shaft WIP QC from context if available
    ctx_items_by_idx: dict = {}
    if ctx and ctx.get("items"):
        for c in ctx["items"]:
            ctx_items_by_idx[int(c.get("item_index"))] = c

    # Per-item section
    items_html = ""
    for item in record.get("items") or []:
        ctx_item = ctx_items_by_idx.get(int(item.get("item_index", 0))) or {}
        pipe_wip = ctx_item.get("pipe_wip_qc")
        shaft_wip = ctx_item.get("shaft_wip_qc")

        # WIP QC table HTML
        wip_html = ""
        if pipe_wip and pipe_wip.get("samples"):
            pipe_rows = ""
            for ps in pipe_wip["samples"]:
                dia_text = "PASS" if ps.get("pipe_dia_ok") is True else ("FAIL" if ps.get("pipe_dia_ok") is False else "–")
                dia_c = "#059669" if ps.get("pipe_dia_ok") is True else ("#DC2626" if ps.get("pipe_dia_ok") is False else "#94A3B8")
                pipe_rows += f"""<tr>
                    <td style="text-align:center;font-weight:700">{ps.get('sample_no','-')}</td>
                    <td style="text-align:center;color:{dia_c};font-weight:700">{dia_text}</td>
                    <td style="text-align:center">{ps.get('pipe_length_measured','–')}</td>
                    <td style="text-align:center">{ps.get('pipe_thickness_measured','–')}</td>
                </tr>"""
            wip_html += f"""<div class="wip-block" style="background:#F0F9FF;border-left:3px solid #0369A1">
              <div class="wip-title" style="color:#0369A1">Pipe WIP QC Results</div>
              <table class="wip-tbl"><thead><tr>
                <th>#</th><th>Dia OK?</th><th>Length (mm)</th><th>Thickness (mm)</th>
              </tr></thead><tbody>{pipe_rows}</tbody></table>
            </div>"""
        if shaft_wip and shaft_wip.get("samples"):
            shaft_rows = ""
            for ss in shaft_wip["samples"]:
                def fmt(v):
                    if v is True: return '<span style="color:#059669;font-weight:700">YES</span>'
                    if v is False: return '<span style="color:#DC2626;font-weight:700">NO</span>'
                    return "–"
                shaft_rows += f"""<tr>
                    <td style="text-align:center;font-weight:700">{ss.get('sample_no','-')}</td>
                    <td style="text-align:center">{fmt(ss.get('length_ok'))}</td>
                    <td style="text-align:center">{fmt(ss.get('width_ok'))}</td>
                    <td style="text-align:center">{fmt(ss.get('dim_ok'))}</td>
                    <td style="text-align:center">{fmt(ss.get('third_ok'))}</td>
                </tr>"""
            wip_html += f"""<div class="wip-block" style="background:#ECFEFF;border-left:3px solid #155E75">
              <div class="wip-title" style="color:#155E75">Shaft WIP QC Results</div>
              <table class="wip-tbl"><thead><tr>
                <th>#</th><th>Length OK?</th><th>Width OK?</th><th>Dim OK?</th><th>3rd OK?</th>
              </tr></thead><tbody>{shaft_rows}</tbody></table>
            </div>"""

        samples = item.get("samples") or []
        rows = ""
        for s in samples:
            overall = "PASS" if s.get("overall_pass") else "FAIL"
            overall_c = "#059669" if s.get("overall_pass") else "#DC2626"
            bearing_reason_html = (
                f"<br><span style=\"font-size:9px;color:#B91C1C\">{s.get('bearing_reason') or ''}</span>"
                if s.get("bearing_match") is False else ""
            )
            rust_reason_html = (
                f"<br><span style=\"font-size:9px;color:#B91C1C\">{s.get('rust_reason') or ''}</span>"
                if s.get("rust_preventive") is False else ""
            )
            cells = [f'<td style="text-align:center;font-weight:700">{s.get("sample_no","-")}</td>']
            if show_runout:
                cells.append(f'<td style="text-align:center">{_fmt(s.get("runout_mm")," mm")}<br>{_pf(s.get("runout_ok"))}</td>')
            if show_water:
                cells.append(f'<td style="text-align:center">{_pf(s.get("water_ok"))}</td>')
            if show_dust:
                cells.append(f'<td style="text-align:center">{_pf(s.get("dust_ok"))}</td>')
            if show_friction:
                cells.append(f'<td style="text-align:center">{_fmt(s.get("friction_coeff"))}<br>{_pf(s.get("friction_ok"))}</td>')
            if show_painting:
                cells.append(f'<td style="text-align:center">{_pf(s.get("painting_visual_ok"))}</td>')
                cells.append(f'<td style="text-align:center">{_fmt(s.get("dft_microns")," µm")}<br>{_pf(s.get("dft_ok"))}</td>')
            cells.append(f'<td style="text-align:center">{_yn(s.get("bearing_match"))}{bearing_reason_html}</td>')
            cells.append(f'<td style="text-align:center">{_yn(s.get("rust_preventive"))}{rust_reason_html}</td>')
            cells.append(f'<td style="text-align:center">{_pf(s.get("welding_ok"))}</td>')
            cells.append(f'<td style="text-align:center;color:{overall_c};font-weight:700">{overall}</td>')
            rows += f"<tr>{''.join(cells)}</tr>"

        tol_html = f'Runout Tol: <b>&lt; {item.get("runout_tolerance_mm","-")} mm</b> &nbsp;·&nbsp; ' if show_runout else ''
        dft_html = f'Expected DFT: <b>{item.get("expected_dft_microns") or "-"} µm</b> (±20%) &nbsp;·&nbsp; ' if show_painting else ''
        items_html += f"""
        <div class="item-block">
          <div class="item-head">
            <div><b>Item #{item.get('item_index')+1}: {item.get('product_name','')}</b> <span style="color:#64748B">({item.get('product_code','')})</span></div>
            <div style="font-size:10px;color:#64748B">
              Qty: <b>{item.get('quantity','-')}</b> &nbsp;·&nbsp;
              Pipe L: <b>{item.get('pipe_length_mm','-')} mm</b> &nbsp;·&nbsp;
              {tol_html}
              {dft_html}
              Bearing (WO): <b>{item.get('bearing_spec','-')}</b>
            </div>
          </div>
          {wip_html}
          <div class="fi-title">Final Inspection Samples</div>
          <table>
            <thead><tr>{head_html}</tr></thead>
            <tbody>{rows if rows else '<tr><td colspan="' + str(len(head_cols)) + '" style="text-align:center;color:#94A3B8;padding:16px">No samples captured</td></tr>'}</tbody>
          </table>
          <div style="font-size:10px;color:#64748B;margin-top:4px">
            Samples: {len(samples)} &nbsp;·&nbsp; Pass: <b style="color:#059669">{item.get('pass_count',0)}</b> &nbsp;·&nbsp; Fail: <b style="color:#DC2626">{item.get('fail_count',0)}</b>
          </div>
        </div>"""

    na_tests = [k for k, v in applicable.items() if not v]
    na_strip = ""
    if na_tests:
        labels = {"runout": "Runout", "water": "Water", "dust": "Dust", "friction": "Friction", "painting": "Painting"}
        na_strip = f'<div style="font-size:10px;color:#92400E;background:#FEF3C7;padding:6px 10px;border-radius:6px;margin:8px 0"><b>Not Applicable per SO:</b> {", ".join(labels[k] for k in na_tests)}</div>'

    inspector = record.get("inspector_name") or record.get("inspected_by") or "-"
    inspected_at = record.get("inspected_at") or ""
    inspected_date = format_date_dmy(inspected_at.split("T")[0]) if inspected_at else "-"

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>Final Inspection Report — {record.get('wo_number','')}</title>
<style>
@page {{ size: A4; margin: 10mm; }}
body {{ font-family: 'Helvetica', sans-serif; margin:0; padding:0; color:#0F172A; position:relative; }}
.page {{ padding: 6mm 4mm; }}
.banner {{ background: #0F172A; color: #fff; padding: 14px 18px; border-radius: 6px; position:relative; }}
.banner h1 {{ margin: 0; font-size: 20px; letter-spacing: 1.5px; }}
.banner .sub {{ font-size: 11px; color:#C5964A; letter-spacing: 0.5px; margin-top:2px }}
.banner .num {{ position: absolute; right: 18px; top: 16px; font-size: 13px; background: rgba(197,150,74,0.2); color:#C5964A; padding: 6px 14px; border-radius: 20px; font-weight: 700; letter-spacing: 1px; }}
.meta-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; margin: 12px 0; }}
.meta {{ background: #F8FAFC; border-left: 3px solid #C5964A; padding: 7px 10px; border-radius: 4px; }}
.meta .k {{ font-size: 8px; color: #64748B; letter-spacing: 0.5px; text-transform: uppercase; }}
.meta .v {{ font-size: 12px; font-weight: 700; color: #0F172A; margin-top: 2px; }}
.item-block {{ margin-top: 12px; border:1px solid #E2E8F0; border-radius:6px; padding:10px; background:#FFFFFF; page-break-inside: avoid; }}
.item-head {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:6px; padding-bottom:6px; border-bottom:1px dashed #CBD5E1; font-size:12px; }}
.wip-block {{ border-radius: 4px; padding: 6px 8px; margin: 8px 0; }}
.wip-title {{ font-size: 10px; font-weight: 800; margin-bottom: 4px; letter-spacing: 0.5px; text-transform: uppercase; }}
.wip-tbl {{ width: 100%; border-collapse: collapse; font-size: 9px; }}
.wip-tbl th {{ background: #E0F2FE; color: #0F172A; padding: 3px 4px; border: 1px solid #BAE6FD; font-weight: 700; }}
.wip-tbl td {{ padding: 3px 4px; border: 1px solid #E0F2FE; background: #FFFFFF; }}
.fi-title {{ font-size: 10px; font-weight: 800; color: #C5964A; margin: 10px 0 4px; letter-spacing: 0.5px; text-transform: uppercase; }}
table {{ width:100%; border-collapse: collapse; font-size: 10px; }}
th {{ padding:5px 4px; background:#0F172A; color:#fff; font-weight:700; border:1px solid #fff; }}
td {{ padding:5px 4px; border:1px solid #CBD5E1; vertical-align: middle; }}
.stamp {{ position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-18deg); font-size: 120px; font-weight: 900; color: {stamp_color}; opacity: 0.10; letter-spacing: 18px; pointer-events: none; }}
.sig-row {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-top: 32px; padding-top: 14px; }}
.sig {{ border-top: 2px dashed #94A3B8; padding-top: 6px; font-size: 11px; text-align: center; color: #475569; font-weight: 600; }}
.sig .sub {{ font-size: 9px; color:#94A3B8; font-weight:500; margin-top:2px }}
.footer {{ font-size: 9px; color: #94A3B8; text-align: center; margin-top: 16px; padding-top: 6px; border-top: 1px solid #E2E8F0; }}
</style></head><body>
<div class="stamp">{status}</div>
<div class="page">
  <div class="banner">
    <h1>FINAL INSPECTION REPORT</h1>
    <div class="sub">Consolidated Post-Assembly QC · Signed Dispatch Clearance</div>
    <div class="num">{record.get('wo_number','')}</div>
  </div>
  <div class="meta-grid">
    <div class="meta"><div class="k">Work Order</div><div class="v">{record.get('wo_number','-')}</div></div>
    <div class="meta"><div class="k">Sales Order</div><div class="v">{wo.get('so_number','-')}</div></div>
    <div class="meta"><div class="k">Customer</div><div class="v">{wo.get('customer_name','-')}</div></div>
    <div class="meta"><div class="k">Delivery Date</div><div class="v">{format_date_dmy(str(wo.get('delivery_date','')).split('T')[0]) if wo.get('delivery_date') else '-'}</div></div>
  </div>
  {na_strip}
  {items_html if items_html else '<div style="padding:20px;text-align:center;color:#94A3B8;border:1px dashed #E2E8F0;border-radius:6px;margin-top:12px">No items recorded yet.</div>'}
  <div class="sig-row">
    <div class="sig">Quality Inspector<div class="sub">{inspector} · {inspected_date}</div></div>
    <div class="sig">Production Head<div class="sub">&nbsp;</div></div>
    <div class="sig">Authorised Signatory<div class="sub">&nbsp;</div></div>
  </div>
  <div class="footer">{COMPANY['name']} · {COMPANY['email']} · {COMPANY['phone']} · Generated {format_date_dmy(inspected_at.split('T')[0]) if inspected_at else '-'}</div>
</div></body></html>"""
    return html


def _verify_token(token: Optional[str], authorization: Optional[str]) -> str:
    auth_token = token
    if not auth_token and authorization and authorization.startswith("Bearer "):
        auth_token = authorization[7:]
    if not auth_token:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        if not payload.get("sub"):
            raise HTTPException(status_code=401, detail="Invalid token")
        return payload.get("sub")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


@router.get("/work-orders/{wo_id}/final-inspection/pdf")
async def final_inspection_pdf(
    wo_id: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    _verify_token(token, authorization)
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work Order not found")
    record = await db.final_inspection_records.find_one({"wo_id": wo_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="No Final Inspection record for this WO")
    ctx = await _build_wo_context(wo)
    html = _render_final_inspection_html(record, wo, ctx)
    buf = io.BytesIO(html.encode("utf-8"))
    buf.seek(0)
    filename = f"{record.get('wo_number','wo').replace('/', '-')}-FinalInspection.html"
    return StreamingResponse(
        buf,
        media_type="text/html",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/work-orders/{wo_id}/final-inspection/excel")
async def final_inspection_excel(
    wo_id: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    _verify_token(token, authorization)
    wo = await db.work_orders.find_one({"id": wo_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Work Order not found")
    record = await db.final_inspection_records.find_one({"wo_id": wo_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="No Final Inspection record for this WO")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Final Inspection"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Meta rows
    ws.append([f"FINAL INSPECTION — {record.get('wo_number','')}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Customer: {wo.get('customer_name','-')}", f"SO: {wo.get('so_number','-')}", f"Status: {(record.get('status') or '').upper()}", f"Inspector: {record.get('inspector_name') or record.get('inspected_by','-')}"])
    ws.append([])

    applicable = record.get("applicable_tests") or {"runout": True, "water": True, "dust": True, "friction": True, "painting": True}
    show_runout = bool(applicable.get("runout", True))
    show_water = bool(applicable.get("water", True))
    show_dust = bool(applicable.get("dust", True))
    show_friction = bool(applicable.get("friction", True))
    show_painting = bool(applicable.get("painting", True))

    headers = ["Item #", "Product", "Code", "Sample #"]
    if show_runout: headers += ["Runout (mm)", "Runout Tol (mm)", "Runout OK?"]
    if show_water: headers += ["Water"]
    if show_dust: headers += ["Dust"]
    if show_friction: headers += ["Friction COF", "Friction OK?"]
    if show_painting: headers += ["Paint Visual", "DFT (µm)", "Expected DFT (µm)", "DFT OK?"]
    headers += ["Bearing Match", "Bearing Reason", "Rust Preventive", "Rust Reason", "Welding", "Overall", "Remarks"]

    ws.append(headers)
    hdr_row = ws.max_row
    for col_i, _ in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=col_i)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for item in record.get("items") or []:
        for s in item.get("samples") or []:
            row: list = [
                item.get("item_index", 0) + 1,
                item.get("product_name", ""),
                item.get("product_code", ""),
                s.get("sample_no", ""),
            ]
            if show_runout:
                row += [s.get("runout_mm"), item.get("runout_tolerance_mm"), ("PASS" if s.get("runout_ok") else "FAIL")]
            if show_water:
                row += [("PASS" if s.get("water_ok") else "FAIL")]
            if show_dust:
                row += [("PASS" if s.get("dust_ok") else "FAIL")]
            if show_friction:
                row += [s.get("friction_coeff"), ("PASS" if s.get("friction_ok") else "FAIL")]
            if show_painting:
                row += [
                    ("OK" if s.get("painting_visual_ok") else "NOT OK"),
                    s.get("dft_microns"),
                    item.get("expected_dft_microns"),
                    ("PASS" if s.get("dft_ok") else "FAIL"),
                ]
            row += [
                "Yes" if s.get("bearing_match") else "No",
                s.get("bearing_reason", "") or "",
                "Yes" if s.get("rust_preventive") else "No",
                s.get("rust_reason", "") or "",
                "OK" if s.get("welding_ok") else "NOT OK",
                "PASS" if s.get("overall_pass") else "FAIL",
                s.get("remarks", "") or "",
            ]
            ws.append(row)

    for col_letter in "ABCDEFGHIJKLMNOPQRSTUV":
        ws.column_dimensions[col_letter].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{record.get('wo_number','wo').replace('/', '-')}-FinalInspection.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
