"""Admin Routes — Raw Material Prices, Standards Data, Drawing Generator"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Body
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
from bson import ObjectId
from routes import db, get_current_user, require_role, ROOT_DIR, UserRole, get_password_hash, generate_customer_code
from routes.price_history import log_price_change
import roller_standards as rs
import io
import base64
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

# ============= ADMIN API - RAW MATERIAL PRICES =============

class PriceUpdateRequest(BaseModel):
    category: str  # bearing, seal, circlip, pipe, shaft, rubber_ring, locking_ring
    key: str  # e.g., "6204", "20", "89/140"
    sub_key: Optional[str] = None  # e.g., "china", "skf" for bearings; "A", "B", "C" for pipe weight
    value: float

@router.get("/admin/prices")
async def get_all_prices(current_user: dict = Depends(get_current_user)):
    """Get all raw material prices for admin panel"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if there are custom prices in database
    custom_prices = await db.custom_prices.find_one({"_id": "prices"})
    
    # If no custom prices, check for saved defaults
    if not custom_prices:
        default_prices = await db.default_prices.find_one({"_id": "defaults"})
        if default_prices and "prices" in default_prices:
            custom_prices = default_prices["prices"]
    
    # Build the response with current prices (from DB or defaults)
    prices = {
        "basic_rates": {
            "pipe_cost_per_kg": custom_prices.get("pipe_cost_per_kg", rs.PIPE_COST_PER_KG) if custom_prices else rs.PIPE_COST_PER_KG,
            "shaft_cost_per_kg": custom_prices.get("shaft_cost_per_kg", rs.SHAFT_COST_PER_KG) if custom_prices else rs.SHAFT_COST_PER_KG,
        },
        "bearing_costs": custom_prices.get("bearing_costs", rs.BEARING_COSTS) if custom_prices else rs.BEARING_COSTS,
        "housing_costs": custom_prices.get("housing_costs", rs.HOUSING_COSTS) if custom_prices else rs.HOUSING_COSTS,
        "seal_costs": custom_prices.get("seal_costs", rs.SEAL_COSTS) if custom_prices else rs.SEAL_COSTS,
        "circlip_costs": custom_prices.get("circlip_costs", rs.CIRCLIP_COSTS) if custom_prices else rs.CIRCLIP_COSTS,
        "rubber_ring_costs": custom_prices.get("rubber_ring_costs", rs.RUBBER_RING_COSTS) if custom_prices else rs.RUBBER_RING_COSTS,
        "locking_ring_costs": custom_prices.get("locking_ring_costs", rs.LOCKING_RING_COSTS) if custom_prices else rs.LOCKING_RING_COSTS,
        "pipe_weight": custom_prices.get("pipe_weight", rs.PIPE_WEIGHT_PER_METER) if custom_prices else rs.PIPE_WEIGHT_PER_METER,
        "shaft_weight": custom_prices.get("shaft_weight", rs.SHAFT_WEIGHT_PER_METER) if custom_prices else rs.SHAFT_WEIGHT_PER_METER,
    }
    
    return prices

@router.post("/admin/prices/update")
async def update_price(request: PriceUpdateRequest, current_user: dict = Depends(get_current_user)):
    """Update a specific raw material price"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get or create custom prices document
    custom_prices = await db.custom_prices.find_one({"_id": "prices"})
    if not custom_prices:
        custom_prices = {"_id": "prices"}

    # Capture old value before mutation (for audit log)
    def _get_old(cat: str, key: str, sub: Optional[str]):
        try:
            if cat == "pipe_cost":
                return custom_prices.get("pipe_cost_per_kg", rs.PIPE_COST_PER_KG)
            if cat == "shaft_cost":
                return custom_prices.get("shaft_cost_per_kg", rs.SHAFT_COST_PER_KG)
            if cat == "bearing":
                return (custom_prices.get("bearing_costs") or rs.BEARING_COSTS).get(key, {}).get(sub)
            if cat == "seal":
                return (custom_prices.get("seal_costs") or rs.SEAL_COSTS).get(key)
            if cat == "circlip":
                return (custom_prices.get("circlip_costs") or {str(k): v for k, v in rs.CIRCLIP_COSTS.items()}).get(key)
            if cat == "rubber_ring":
                return (custom_prices.get("rubber_ring_costs") or rs.RUBBER_RING_COSTS).get(key)
            if cat == "locking_ring":
                return (custom_prices.get("locking_ring_costs") or {str(k): v for k, v in rs.LOCKING_RING_COSTS.items()}).get(key)
            if cat == "housing":
                return (custom_prices.get("housing_costs") or rs.HOUSING_COSTS).get(key)
            if cat == "pipe_weight":
                return (custom_prices.get("pipe_weight") or {str(k): v for k, v in rs.PIPE_WEIGHT_PER_METER.items()}).get(key, {}).get(sub)
            if cat == "shaft_weight":
                return (custom_prices.get("shaft_weight") or {str(k): v for k, v in rs.SHAFT_WEIGHT_PER_METER.items()}).get(key)
        except Exception:
            return None
        return None

    old_value = _get_old(request.category, request.key, request.sub_key)
    
    # Update based on category
    if request.category == "pipe_cost":
        custom_prices["pipe_cost_per_kg"] = request.value
    elif request.category == "shaft_cost":
        custom_prices["shaft_cost_per_kg"] = request.value
    elif request.category == "bearing":
        if "bearing_costs" not in custom_prices:
            import copy
            custom_prices["bearing_costs"] = copy.deepcopy(rs.BEARING_COSTS)
        if request.key not in custom_prices["bearing_costs"]:
            custom_prices["bearing_costs"][request.key] = {}
        custom_prices["bearing_costs"][request.key][request.sub_key] = request.value
    elif request.category == "seal":
        if "seal_costs" not in custom_prices:
            custom_prices["seal_costs"] = dict(rs.SEAL_COSTS)
        custom_prices["seal_costs"][request.key] = request.value
    elif request.category == "circlip":
        if "circlip_costs" not in custom_prices:
            custom_prices["circlip_costs"] = {str(k): v for k, v in rs.CIRCLIP_COSTS.items()}
        custom_prices["circlip_costs"][request.key] = request.value
    elif request.category == "rubber_ring":
        if "rubber_ring_costs" not in custom_prices:
            custom_prices["rubber_ring_costs"] = dict(rs.RUBBER_RING_COSTS)
        custom_prices["rubber_ring_costs"][request.key] = request.value
    elif request.category == "locking_ring":
        if "locking_ring_costs" not in custom_prices:
            custom_prices["locking_ring_costs"] = {str(k): v for k, v in rs.LOCKING_RING_COSTS.items()}
        custom_prices["locking_ring_costs"][request.key] = request.value
    elif request.category == "housing":
        if "housing_costs" not in custom_prices:
            custom_prices["housing_costs"] = dict(rs.HOUSING_COSTS)
        custom_prices["housing_costs"][request.key] = request.value
    elif request.category == "pipe_weight":
        if "pipe_weight" not in custom_prices:
            import copy
            custom_prices["pipe_weight"] = copy.deepcopy({str(k): v for k, v in rs.PIPE_WEIGHT_PER_METER.items()})
        if request.key not in custom_prices["pipe_weight"]:
            custom_prices["pipe_weight"][request.key] = {}
        custom_prices["pipe_weight"][request.key][request.sub_key] = request.value
    elif request.category == "shaft_weight":
        if "shaft_weight" not in custom_prices:
            custom_prices["shaft_weight"] = {str(k): v for k, v in rs.SHAFT_WEIGHT_PER_METER.items()}
        custom_prices["shaft_weight"][request.key] = request.value
    else:
        raise HTTPException(status_code=400, detail=f"Unknown category: {request.category}")
    
    custom_prices["updated_at"] = datetime.utcnow().isoformat()
    custom_prices["updated_by"] = current_user.get("email")
    
    # Save to database
    await db.custom_prices.replace_one({"_id": "prices"}, custom_prices, upsert=True)
    
    # Invalidate price cache so calculations use new values immediately
    import price_loader
    price_loader.invalidate_cache()

    # Audit log (fire-and-forget)
    try:
        await log_price_change(
            user_email=current_user.get("email") or "",
            product_type="roller",
            category=request.category,
            key=request.key,
            sub_key=request.sub_key,
            old_value=old_value if isinstance(old_value, (int, float)) else None,
            new_value=float(request.value),
        )
    except Exception:
        pass

    return {"message": "Price updated successfully", "category": request.category, "key": request.key}

# Set as Default - Send OTP for verification
@router.post("/admin/prices/set-default/send-otp")
async def send_set_default_otp(current_user: dict = Depends(get_current_user)):
    """Send OTP to admin email for setting prices as default"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    email = current_user.get("email")
    name = current_user.get("name", "Admin")
    
    # Check cooldown
    existing_otp = await db.price_otp_verifications.find_one({"email": email})
    if existing_otp:
        created_at = existing_otp.get("created_at")
        if created_at:
            # Make sure both datetimes are timezone-aware
            if created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)
            elapsed = (datetime.now(timezone.utc) - created_at).total_seconds()
            if elapsed < OTP_COOLDOWN_SECONDS:
                remaining = int(OTP_COOLDOWN_SECONDS - elapsed)
                raise HTTPException(
                    status_code=429,
                    detail=f"Please wait {remaining} seconds before requesting a new OTP"
                )
    
    # Generate and store OTP
    otp = generate_otp()
    expiry = datetime.now(timezone.utc) + timedelta(minutes=OTP_EXPIRY_MINUTES)
    
    await db.price_otp_verifications.update_one(
        {"email": email},
        {
            "$set": {
                "otp": otp,
                "expires_at": expiry,
                "created_at": datetime.now(timezone.utc),
                "verified": False,
                "purpose": "set_default_prices"
            }
        },
        upsert=True
    )
    
    # Send OTP email
    try:
        if GMAIL_USER and GMAIL_APP_PASSWORD:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f"Price Update Verification Code - {otp}"
            msg['From'] = f"Convero Solutions <{GMAIL_USER}>"
            msg['To'] = email
            
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                    .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                    .header {{ background-color: #960018; color: white; padding: 20px; text-align: center; }}
                    .content {{ padding: 30px; background-color: #f9f9f9; }}
                    .otp-box {{ background-color: #960018; color: white; font-size: 32px; font-weight: bold; letter-spacing: 8px; padding: 20px 40px; text-align: center; border-radius: 8px; margin: 20px 0; }}
                    .warning {{ background-color: #fff3cd; border: 1px solid #ffc107; padding: 15px; border-radius: 8px; margin: 20px 0; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Price Update Verification</h1>
                    </div>
                    <div class="content">
                        <p>Dear {name},</p>
                        <p>You have requested to set current prices as default. Please use the following verification code:</p>
                        <div class="otp-box">{otp}</div>
                        <div class="warning">
                            <strong>⚠️ Warning:</strong> This action will update the default prices in the system. All future calculations will use these new rates.
                        </div>
                        <p>This code will expire in {OTP_EXPIRY_MINUTES} minutes.</p>
                        <p>If you did not request this, please ignore this email.</p>
                    </div>
                </div>
            </body>
            </html>
            """
            
            msg.attach(MIMEText(html_content, 'html'))
            
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                server.sendmail(GMAIL_USER, email, msg.as_string())
            
            logging.info(f"Set default OTP sent to {email}")
        else:
            logging.warning("Email not configured, OTP not sent")
            
    except Exception as e:
        logging.error(f"Failed to send OTP email: {e}")
        # Continue anyway for development
    
    return {"message": f"Verification code sent to {email}", "email": email}

# Set as Default - Verify OTP and update defaults
@router.post("/admin/prices/set-default/verify")
async def verify_and_set_default_prices(
    otp: str = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """Verify OTP and set current prices as new defaults"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    email = current_user.get("email")
    
    # Verify OTP
    otp_record = await db.price_otp_verifications.find_one({"email": email})
    if not otp_record:
        raise HTTPException(status_code=400, detail="No OTP request found. Please request a new code.")
    
    if otp_record.get("otp") != otp:
        raise HTTPException(status_code=400, detail="Invalid verification code")
    
    expires_at = otp_record.get("expires_at")
    if expires_at:
        # Make sure both datetimes are timezone-aware
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if datetime.now(timezone.utc) > expires_at:
            raise HTTPException(status_code=400, detail="Verification code has expired. Please request a new one.")
    
    # Get current custom prices
    custom_prices = await db.custom_prices.find_one({"_id": "prices"})
    
    if not custom_prices:
        raise HTTPException(status_code=400, detail="No custom prices found. Current prices are already defaults.")
    
    # Update roller_standards.py with new default values
    try:
        import roller_standards as rs
        
        # Update basic rates
        if "pipe_cost_per_kg" in custom_prices:
            rs.PIPE_COST_PER_KG = custom_prices["pipe_cost_per_kg"]
        if "shaft_cost_per_kg" in custom_prices:
            rs.SHAFT_COST_PER_KG = custom_prices["shaft_cost_per_kg"]
        
        # Update bearing costs
        if "bearing_costs" in custom_prices:
            for bearing, makes in custom_prices["bearing_costs"].items():
                if bearing in rs.BEARING_COSTS:
                    rs.BEARING_COSTS[bearing].update(makes)
                else:
                    rs.BEARING_COSTS[bearing] = makes
        
        # Update housing costs
        if "housing_costs" in custom_prices:
            rs.HOUSING_COSTS.update(custom_prices["housing_costs"])
        
        # Update seal costs
        if "seal_costs" in custom_prices:
            rs.SEAL_COSTS.update(custom_prices["seal_costs"])
        
        # Update circlip costs
        if "circlip_costs" in custom_prices:
            for shaft, cost in custom_prices["circlip_costs"].items():
                rs.CIRCLIP_COSTS[int(shaft)] = cost
        
        # Update rubber ring costs
        if "rubber_ring_costs" in custom_prices:
            rs.RUBBER_RING_COSTS.update(custom_prices["rubber_ring_costs"])
        
        # Update locking ring costs
        if "locking_ring_costs" in custom_prices:
            for pipe, cost in custom_prices["locking_ring_costs"].items():
                rs.LOCKING_RING_COSTS[int(pipe)] = cost
        
        # Store the update in a permanent collection for persistence across restarts
        await db.default_prices.update_one(
            {"_id": "defaults"},
            {
                "$set": {
                    "prices": custom_prices,
                    "updated_at": datetime.now(timezone.utc),
                    "updated_by": email
                }
            },
            upsert=True
        )
        
        # Clear custom prices since they are now defaults
        await db.custom_prices.delete_one({"_id": "prices"})
        
        # Invalidate price cache
        import price_loader
        price_loader.invalidate_cache()
        
        # Delete OTP record
        await db.price_otp_verifications.delete_one({"email": email})
        
        logging.info(f"Default prices updated by {email}")
        
        return {
            "message": "Prices have been set as new defaults successfully!",
            "updated_by": email,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
    except Exception as e:
        logging.error(f"Failed to set default prices: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update defaults: {str(e)}")

@router.post("/admin/prices/reset")
async def reset_prices(current_user: dict = Depends(get_current_user)):
    """Reset all prices to default values"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await db.custom_prices.delete_one({"_id": "prices"})
    
    # Invalidate price cache so calculations use default values immediately
    import price_loader
    price_loader.invalidate_cache()
    
    return {"message": "All prices reset to default values"}

@router.get("/admin/prices/export")
async def export_prices_to_excel(token: Optional[str] = None):
    """Export all prices to Excel file. Accepts token as query param for browser downloads."""
    # Validate token from query param
    current_user = None
    if token:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            email = payload.get("sub")
            if email:
                user = await db.users.find_one({"email": email})
                if user:
                    current_user = user
        except Exception as e:
            logging.error(f"Token validation error: {e}")
            raise HTTPException(status_code=401, detail="Invalid token")
    
    if not current_user or current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get current prices - check custom_prices first, then default_prices
    custom_prices = await db.custom_prices.find_one({"_id": "prices"})
    if not custom_prices:
        default_prices = await db.default_prices.find_one({"_id": "defaults"})
        if default_prices and "prices" in default_prices:
            custom_prices = default_prices["prices"]
        else:
            custom_prices = {}
    
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="960018", end_color="960018", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Basic Rates
    ws_basic = wb.active
    ws_basic.title = "Basic Rates"
    ws_basic.append(["Item", "Cost (Rs)"])
    ws_basic.append(["Pipe Cost per kg", custom_prices.get("pipe_cost_per_kg", rs.PIPE_COST_PER_KG)])
    ws_basic.append(["Shaft Cost per kg", custom_prices.get("shaft_cost_per_kg", rs.SHAFT_COST_PER_KG)])
    for cell in ws_basic[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_basic.column_dimensions['A'].width = 25
    ws_basic.column_dimensions['B'].width = 15
    
    # Sheet 2: Bearing Costs
    ws_bearing = wb.create_sheet("Bearing Costs")
    ws_bearing.append(["Bearing Type", "Shaft Dia (mm)", "Cost (Rs)"])
    bearing_costs = custom_prices.get("bearing_costs", rs.BEARING_COSTS)
    for bearing_type, shaft_costs in bearing_costs.items():
        for shaft_dia, cost in shaft_costs.items():
            ws_bearing.append([bearing_type, shaft_dia, cost])
    for cell in ws_bearing[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_bearing.column_dimensions['A'].width = 15
    ws_bearing.column_dimensions['B'].width = 15
    ws_bearing.column_dimensions['C'].width = 15
    
    # Sheet 3: Housing Costs
    ws_housing = wb.create_sheet("Housing Costs")
    ws_housing.append(["Housing Config (OD/Bearing)", "Cost (Rs)"])
    housing_costs = custom_prices.get("housing_costs", rs.HOUSING_COSTS)
    for config, cost in housing_costs.items():
        ws_housing.append([config, cost])
    for cell in ws_housing[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_housing.column_dimensions['A'].width = 25
    ws_housing.column_dimensions['B'].width = 15
    
    # Sheet 4: Seal Costs
    ws_seal = wb.create_sheet("Seal Costs")
    ws_seal.append(["Seal Type", "Cost (Rs)"])
    seal_costs = custom_prices.get("seal_costs", rs.SEAL_COSTS)
    for seal_type, cost in seal_costs.items():
        ws_seal.append([seal_type, cost])
    for cell in ws_seal[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_seal.column_dimensions['A'].width = 15
    ws_seal.column_dimensions['B'].width = 15
    
    # Sheet 5: Circlip Costs
    ws_circlip = wb.create_sheet("Circlip Costs")
    ws_circlip.append(["Shaft Dia (mm)", "Cost (Rs)"])
    circlip_costs = custom_prices.get("circlip_costs", rs.CIRCLIP_COSTS)
    for shaft, cost in circlip_costs.items():
        ws_circlip.append([shaft, cost])
    for cell in ws_circlip[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_circlip.column_dimensions['A'].width = 15
    ws_circlip.column_dimensions['B'].width = 15
    
    # Sheet 6: Rubber Ring Costs
    ws_rubber = wb.create_sheet("Rubber Ring Costs")
    ws_rubber.append(["Pipe/Rubber Config", "Cost (Rs)"])
    rubber_costs = custom_prices.get("rubber_ring_costs", rs.RUBBER_RING_COSTS)
    for config, cost in rubber_costs.items():
        ws_rubber.append([config, cost])
    for cell in ws_rubber[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_rubber.column_dimensions['A'].width = 25
    ws_rubber.column_dimensions['B'].width = 15
    
    # Sheet 7: Locking Ring Costs
    ws_locking = wb.create_sheet("Locking Ring Costs")
    ws_locking.append(["Pipe Dia (mm)", "Cost (Rs)"])
    locking_costs = custom_prices.get("locking_ring_costs", rs.LOCKING_RING_COSTS)
    for pipe, cost in locking_costs.items():
        ws_locking.append([pipe, cost])
    for cell in ws_locking[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_locking.column_dimensions['A'].width = 15
    ws_locking.column_dimensions['B'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"convero_prices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/admin/prices/export/pdf")
async def export_prices_to_pdf(token: Optional[str] = None):
    """Export all prices to PDF file"""
    current_user = None
    if token:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            email = payload.get("sub")
            if email:
                user = await db.users.find_one({"email": email})
                if user:
                    current_user = user
        except Exception as e:
            logging.error(f"Token validation error: {e}")
            raise HTTPException(status_code=401, detail="Invalid token")
    
    if not current_user or current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get current prices - check custom_prices first, then default_prices
    custom_prices = await db.custom_prices.find_one({"_id": "prices"})
    if not custom_prices:
        default_prices = await db.default_prices.find_one({"_id": "defaults"})
        if default_prices and "prices" in default_prices:
            custom_prices = default_prices["prices"]
        else:
            custom_prices = {}
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Price List</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; font-size: 11px; }}
            h1 {{ color: #960018; border-bottom: 2px solid #960018; padding-bottom: 10px; font-size: 24px; }}
            h2 {{ color: #960018; margin-top: 20px; font-size: 16px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            th {{ background-color: #960018; color: white; padding: 8px; text-align: left; }}
            td {{ padding: 6px; border-bottom: 1px solid #ddd; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
            .two-col {{ display: flex; gap: 20px; }}
            .two-col > div {{ flex: 1; }}
            .footer {{ margin-top: 30px; text-align: center; color: #666; font-size: 10px; }}
        </style>
    </head>
    <body>
        <h1>Convero Price List</h1>
        <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <h2>Basic Rates</h2>
        <table>
            <tr><th>Item</th><th>Cost (Rs)</th></tr>
            <tr><td>Pipe Cost per kg</td><td>{custom_prices.get("pipe_cost_per_kg", rs.PIPE_COST_PER_KG)}</td></tr>
            <tr><td>Shaft Cost per kg</td><td>{custom_prices.get("shaft_cost_per_kg", rs.SHAFT_COST_PER_KG)}</td></tr>
        </table>
        
        <div class="two-col">
            <div>
                <h2>Housing Costs</h2>
                <table>
                    <tr><th>Config</th><th>Cost (Rs)</th></tr>
    """
    
    housing_costs = custom_prices.get("housing_costs", rs.HOUSING_COSTS)
    for config, cost in list(housing_costs.items())[:10]:
        html_content += f"<tr><td>{config}</td><td>{cost}</td></tr>"
    
    html_content += """
                </table>
            </div>
            <div>
                <h2>Seal Costs</h2>
                <table>
                    <tr><th>Type</th><th>Cost (Rs)</th></tr>
    """
    
    seal_costs = custom_prices.get("seal_costs", rs.SEAL_COSTS)
    for seal_type, cost in seal_costs.items():
        html_content += f"<tr><td>{seal_type}</td><td>{cost}</td></tr>"
    
    html_content += """
                </table>
            </div>
        </div>
        
        <div class="footer">
            <p>Convero - Belt Conveyor Roller Solutions</p>
        </div>
    </body>
    </html>
    """
    
    from weasyprint import HTML
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    filename = f"convero_prices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/admin/prices/import")
async def import_prices_from_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import prices from Excel file - supports user's actual Excel format"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        sheets = wb.worksheets
        
        # Log available sheet names for debugging
        logger.info(f"[Import] Excel has {len(sheets)} sheets: {wb.sheetnames}")
        
        # Get existing custom prices or create new
        custom_prices = await db.custom_prices.find_one({"_id": "prices"}) or {"_id": "prices"}
        
        updates = {"basic": 0, "pipe": 0, "shaft": 0, "bearing": 0, "housing": 0, "seal": 0, "circlip": 0, "rubber": 0, "locking": 0}
        
        # Process sheets by index (since sheets are unnamed)
        # Sheet 1 (index 0): Pipe data - Pipe OD, Type A/B/C weights, Cost
        if len(sheets) > 0:
            ws = sheets[0]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 1 headers: {headers}")
            
            if "pipe_weights" not in custom_prices:
                custom_prices["pipe_weights"] = {}
            
            # Check if this is pipe data (has Pipe OD and Cost columns)
            if any("pipe" in str(h).lower() for h in headers if h) and any("cost" in str(h).lower() for h in headers if h):
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:
                        pipe_od = str(row[0])
                        # Get pipe cost (usually last column)
                        cost_col = len(row) - 1
                        if row[cost_col] is not None:
                            try:
                                cost = float(row[cost_col])
                                custom_prices["pipe_cost_per_kg"] = cost  # Set global pipe cost
                                updates["pipe"] += 1
                            except (ValueError, TypeError):
                                pass
                        
                        # Store pipe weights by type
                        weights = {}
                        if len(row) > 1 and row[1] is not None:
                            try:
                                weights["type_a"] = float(row[1])
                            except:
                                pass
                        if len(row) > 2 and row[2] is not None:
                            try:
                                weights["type_b"] = float(row[2])
                            except:
                                pass
                        if len(row) > 3 and row[3] is not None:
                            try:
                                weights["type_c"] = float(row[3])
                            except:
                                pass
                        if weights:
                            custom_prices["pipe_weights"][pipe_od] = weights
        
        # Sheet 2 (index 1): Shaft data - Shaft Dia, Weight, Cost
        if len(sheets) > 1:
            ws = sheets[1]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 2 headers: {headers}")
            
            if any("shaft" in str(h).lower() for h in headers if h):
                if "shaft_weights" not in custom_prices:
                    custom_prices["shaft_weights"] = {}
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:
                        shaft_dia = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0])
                        
                        # Weight in column 2
                        if len(row) > 1 and row[1] is not None:
                            try:
                                weight = float(row[1])
                                custom_prices["shaft_weights"][shaft_dia] = weight
                            except:
                                pass
                        
                        # Cost in column 3
                        if len(row) > 2 and row[2] is not None:
                            try:
                                cost = float(row[2])
                                custom_prices["shaft_cost_per_kg"] = cost
                                updates["shaft"] += 1
                            except:
                                pass
        
        # Sheet 3 (index 2): Bearing data - Bearing No, Shaft, China/SKF/FAG/Timken prices
        if len(sheets) > 2:
            ws = sheets[2]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 3 headers: {headers}")
            
            if any("bearing" in str(h).lower() for h in headers if h):
                if "bearing_costs" not in custom_prices:
                    custom_prices["bearing_costs"] = {}
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None and len(row) > 2:
                        bearing_no = str(row[0])
                        
                        # Parse bearing prices by brand (columns: Bearing No, Shaft, China, SKF, FAG, Timken)
                        brands = ["China", "SKF", "FAG", "Timken"]
                        for i, brand in enumerate(brands):
                            col_idx = i + 2  # Start from column 3 (index 2)
                            if len(row) > col_idx and row[col_idx] is not None:
                                try:
                                    val = row[col_idx]
                                    if val and str(val) != '-':
                                        cost = float(val)
                                        if brand not in custom_prices["bearing_costs"]:
                                            custom_prices["bearing_costs"][brand] = {}
                                        custom_prices["bearing_costs"][brand][bearing_no] = cost
                                        updates["bearing"] += 1
                                except (ValueError, TypeError):
                                    pass
        
        # Sheet 4 (index 3): Housing data - Housing/Bore, Price
        if len(sheets) > 3:
            ws = sheets[3]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 4 headers: {headers}")
            
            if "housing_costs" not in custom_prices:
                custom_prices["housing_costs"] = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and len(row) > 1 and row[1] is not None:
                    config = str(row[0])  # e.g., "56/47"
                    try:
                        cost = float(row[1])
                        custom_prices["housing_costs"][config] = cost
                        updates["housing"] += 1
                    except (ValueError, TypeError):
                        pass
        
        # Sheet 5 (index 4): Seal data - Bearing No, Seal Cost (Rs/set)
        # Headers: ['Bearing No', 'Seal Cost (Rs/set)'] - only 2 columns
        if len(sheets) > 4:
            ws = sheets[4]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 5 headers: {headers}")
            
            if "seal_costs" not in custom_prices:
                custom_prices["seal_costs"] = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and len(row) > 1 and row[1] is not None:
                    bearing_no = str(row[0])
                    # Cost is in column 2 (index 1) - Seal Cost (Rs/set)
                    try:
                        cost = float(row[1])
                        custom_prices["seal_costs"][bearing_no] = cost
                        updates["seal"] += 1
                        logger.info(f"[Import] Seal cost: {bearing_no} = {cost}")
                    except (ValueError, TypeError):
                        pass
        
        # Sheet 6 (index 5): Circlip data - Shaft Dia, Price, Qty
        if len(sheets) > 5:
            ws = sheets[5]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 6 headers: {headers}")
            
            if "circlip_costs" not in custom_prices:
                custom_prices["circlip_costs"] = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and len(row) > 1 and row[1] is not None:
                    shaft = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0])
                    try:
                        cost = float(row[1])
                        custom_prices["circlip_costs"][shaft] = cost
                        updates["circlip"] += 1
                    except (ValueError, TypeError):
                        pass
        
        # Sheet 7 (index 6): Rubber Ring data - Pipe/Rubber, Weight, Price
        if len(sheets) > 6:
            ws = sheets[6]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 7 headers: {headers}")
            
            if "rubber_ring_costs" not in custom_prices:
                custom_prices["rubber_ring_costs"] = {}
            if "rubber_ring_weights" not in custom_prices:
                custom_prices["rubber_ring_weights"] = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    config = str(row[0])  # e.g., "60/90"
                    
                    # Weight in column 2
                    if len(row) > 1 and row[1] is not None:
                        try:
                            weight = float(row[1])
                            custom_prices["rubber_ring_weights"][config] = weight
                        except:
                            pass
                    
                    # Price in column 3
                    if len(row) > 2 and row[2] is not None:
                        try:
                            cost = float(row[2])
                            custom_prices["rubber_ring_costs"][config] = cost
                            updates["rubber"] += 1
                        except (ValueError, TypeError):
                            pass
        
        # Sheet 8 (index 7): Locking Ring data - Pipe OD, Price
        if len(sheets) > 7:
            ws = sheets[7]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 8 headers: {headers}")
            
            if "locking_ring_costs" not in custom_prices:
                custom_prices["locking_ring_costs"] = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and len(row) > 1 and row[1] is not None:
                    pipe = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0])
                    try:
                        cost = float(row[1])
                        custom_prices["locking_ring_costs"][pipe] = cost
                        updates["locking"] += 1
                    except (ValueError, TypeError):
                        pass
        
        # Sheet 9 (index 8): Basic Rates - Description, Value (Pipe cost, Shaft cost, etc.)
        if len(sheets) > 8:
            ws = sheets[8]
            headers = [cell.value for cell in ws[1]] if ws[1] else []
            logger.info(f"[Import] Sheet 9 headers: {headers}")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and len(row) > 1 and row[1] is not None:
                    desc = str(row[0]).lower()
                    value_str = str(row[1])
                    
                    try:
                        # First try to parse the value directly as a number
                        try:
                            value = float(row[1])
                        except (ValueError, TypeError):
                            # Extract numeric value from strings like "Rs.67" or "Rs 67" or "67"
                            import re
                            # Remove all non-numeric characters except dots
                            clean_str = re.sub(r'[^\d.]', '', value_str)
                            # Remove leading dots (e.g., ".67" -> "67")
                            clean_str = clean_str.lstrip('.')
                            if clean_str:
                                value = float(clean_str)
                            else:
                                continue
                        
                        logger.info(f"[Import] Parsed '{value_str}' -> {value}")
                        
                        if "pipe" in desc and "cost" in desc:
                            custom_prices["pipe_cost_per_kg"] = value
                            updates["basic"] += 1
                            logger.info(f"[Import] Set pipe cost: {value}")
                        elif "shaft" in desc and "cost" in desc:
                            custom_prices["shaft_cost_per_kg"] = value
                            updates["basic"] += 1
                            logger.info(f"[Import] Set shaft cost: {value}")
                        elif "manufacturing" in desc and "margin" in desc:
                            custom_prices["manufacturing_margin"] = value
                            updates["basic"] += 1
                        elif "overhead" in desc:
                            custom_prices["overhead_factor"] = value
                            updates["basic"] += 1
                    except (ValueError, TypeError):
                        pass
        
        # Save to database
        logger.info(f"[Import] Saving custom_prices with keys: {list(custom_prices.keys())}")
        await db.custom_prices.replace_one({"_id": "prices"}, custom_prices, upsert=True)
        
        # Invalidate price cache
        import price_loader
        price_loader.invalidate_cache()
        
        total_updates = sum(updates.values())
        
        return {
            "message": f"Successfully imported {total_updates} price entries",
            "details": updates
        }
        
    except Exception as e:
        logging.error(f"Error importing prices: {e}")
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {str(e)}")

@router.post("/admin/make-admin")
async def make_user_admin(email: str, current_user: dict = Depends(get_current_user)):
    """Make a user an admin (only existing admins can do this)"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.users.update_one(
        {"email": email},
        {"$set": {"role": UserRole.ADMIN}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": f"User {email} is now an admin"}

# ============= ADMIN API - STANDARDS DATA (MongoDB) =============

@router.get("/admin/standards/{collection}")
async def get_standards_data(collection: str, current_user: dict = Depends(get_current_user)):
    """Get all documents from a standards collection"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    valid_collections = [
        "pipe_diameters", "shaft_diameters", "shaft_end_types", "bearings", 
        "housings", "pipe_weights", "roller_lengths", "circlips", 
        "rubber_lagging", "rubber_rings", "locking_rings", "discount_slabs",
        "freight_rates", "packing_options", "gst_config", "raw_material_costs"
    ]
    
    if collection not in valid_collections:
        raise HTTPException(status_code=400, detail=f"Invalid collection. Valid: {valid_collections}")
    
    cursor = db[collection].find({}, {"_id": 0})
    docs = await cursor.to_list(length=500)
    return {"collection": collection, "count": len(docs), "data": docs}

@router.post("/admin/standards/{collection}")
async def add_standards_item(collection: str, item: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    """Add a new item to a standards collection"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    valid_collections = [
        "pipe_diameters", "shaft_diameters", "shaft_end_types", "bearings", 
        "housings", "pipe_weights", "roller_lengths", "circlips", 
        "rubber_lagging", "rubber_rings", "locking_rings", "discount_slabs",
        "freight_rates", "packing_options", "gst_config", "raw_material_costs"
    ]
    
    if collection not in valid_collections:
        raise HTTPException(status_code=400, detail=f"Invalid collection")
    
    item["created_at"] = datetime.utcnow()
    item["created_by"] = current_user.get("email")
    
    result = await db[collection].insert_one(item)
    return {"message": "Item added successfully", "id": str(result.inserted_id)}

@router.put("/admin/standards/{collection}")
async def update_standards_item(
    collection: str, 
    query: Dict[str, Any],
    update_data: Dict[str, Any],
    current_user: dict = Depends(get_current_user)
):
    """Update an item in a standards collection"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    valid_collections = [
        "pipe_diameters", "shaft_diameters", "shaft_end_types", "bearings", 
        "housings", "pipe_weights", "roller_lengths", "circlips", 
        "rubber_lagging", "rubber_rings", "locking_rings", "discount_slabs",
        "freight_rates", "packing_options", "gst_config", "raw_material_costs"
    ]
    
    if collection not in valid_collections:
        raise HTTPException(status_code=400, detail=f"Invalid collection")
    
    update_data["updated_at"] = datetime.utcnow()
    update_data["updated_by"] = current_user.get("email")
    
    result = await db[collection].update_one(query, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    return {"message": "Item updated successfully", "modified": result.modified_count}

@router.delete("/admin/standards/{collection}")
async def delete_standards_item(
    collection: str,
    query: Dict[str, Any],
    current_user: dict = Depends(get_current_user)
):
    """Delete an item from a standards collection"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    valid_collections = [
        "pipe_diameters", "shaft_diameters", "shaft_end_types", "bearings", 
        "housings", "pipe_weights", "roller_lengths", "circlips", 
        "rubber_lagging", "rubber_rings", "locking_rings", "discount_slabs",
        "freight_rates", "packing_options", "gst_config", "raw_material_costs"
    ]
    
    if collection not in valid_collections:
        raise HTTPException(status_code=400, detail=f"Invalid collection")
    
    result = await db[collection].delete_one(query)
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    return {"message": "Item deleted successfully"}

@router.get("/admin/standards-summary")
async def get_standards_summary(current_user: dict = Depends(get_current_user)):
    """Get a summary of all standards collections"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    collections = [
        "pipe_diameters", "shaft_diameters", "shaft_end_types", "bearings", 
        "housings", "pipe_weights", "roller_lengths", "circlips", 
        "rubber_lagging", "rubber_rings", "locking_rings", "discount_slabs",
        "freight_rates", "packing_options", "gst_config", "raw_material_costs"
    ]
    
    summary = []
    for coll in collections:
        count = await db[coll].count_documents({})
        summary.append({"collection": coll, "count": count})
    
    return {"summary": summary, "total_collections": len(collections)}

# ============= DRAWING GENERATOR =============

class DrawingRequest(BaseModel):
    product_code: str
    roller_type: str
    pipe_diameter: float
    pipe_length: float
    pipe_type: str
    shaft_diameter: float
    bearing: str
    bearing_make: str
    housing: str
    weight_kg: float
    unit_price: float = 0  # Optional, not displayed
    rubber_diameter: Optional[float] = None
    belt_widths: Optional[List[int]] = None
    quantity: int = 1
    shaft_end_type: Optional[str] = "B"  # A (+26mm), B (+36mm), C (+56mm), custom
    custom_shaft_extension: Optional[int] = None  # Custom shaft extension in mm

@router.get("/download/sample-drawing")
async def download_sample_drawing():
    """Download sample roller drawing PDF"""
    file_path = ROOT_DIR / "static" / "sample_drawing.pdf"
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Sample drawing not found")
    return FileResponse(
        path=str(file_path),
        filename="Sample_Roller_Drawing.pdf",
        media_type="application/pdf"
    )

@router.post("/generate-drawing")
async def generate_drawing(request: DrawingRequest, current_user: dict = Depends(get_current_user)):
    """Generate a technical drawing PDF for a roller"""
    from drawing_generator import generate_roller_drawing
    from fastapi.responses import StreamingResponse
    
    try:
        pdf_buffer = generate_roller_drawing(
            product_code=request.product_code,
            roller_type=request.roller_type,
            pipe_diameter=request.pipe_diameter,
            pipe_length=request.pipe_length,
            pipe_type=request.pipe_type,
            shaft_diameter=request.shaft_diameter,
            bearing=request.bearing,
            bearing_make=request.bearing_make,
            housing=request.housing,
            weight_kg=request.weight_kg,
            unit_price=request.unit_price,
            rubber_diameter=request.rubber_diameter,
            belt_widths=request.belt_widths,
            quantity=request.quantity,
            shaft_end_type=request.shaft_end_type or "B",
            custom_shaft_extension=request.custom_shaft_extension
        )
        
        filename = f"Drawing_{request.product_code.replace(' ', '_').replace('/', '-')}.pdf"
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate drawing: {str(e)}")

@router.post("/generate-drawing-base64")
async def generate_drawing_base64(request: DrawingRequest, current_user: dict = Depends(get_current_user)):
    """Generate a technical drawing PDF and return as base64 for mobile apps"""
    from drawing_generator import generate_roller_drawing
    import base64
    
    try:
        pdf_buffer = generate_roller_drawing(
            product_code=request.product_code,
            roller_type=request.roller_type,
            pipe_diameter=request.pipe_diameter,
            pipe_length=request.pipe_length,
            pipe_type=request.pipe_type,
            shaft_diameter=request.shaft_diameter,
            bearing=request.bearing,
            bearing_make=request.bearing_make,
            housing=request.housing,
            weight_kg=request.weight_kg,
            unit_price=request.unit_price,
            rubber_diameter=request.rubber_diameter,
            belt_widths=request.belt_widths,
            quantity=request.quantity,
            shaft_end_type=request.shaft_end_type or "B",
            custom_shaft_extension=request.custom_shaft_extension
        )
        
        # Convert to base64
        pdf_bytes = pdf_buffer.getvalue()
        base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
        
        return {
            "base64": base64_pdf,
            "filename": f"Drawing_{request.product_code.replace(' ', '_').replace('/', '-')}.pdf"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate drawing: {str(e)}")


@router.post("/generate-drawing-download")
async def generate_drawing_download(request: DrawingRequest, current_user: dict = Depends(get_current_user)):
    """Generate and directly download a technical drawing PDF"""
    from drawing_generator import generate_roller_drawing
    from fastapi.responses import Response
    
    try:
        pdf_buffer = generate_roller_drawing(
            product_code=request.product_code,
            roller_type=request.roller_type,
            pipe_diameter=request.pipe_diameter,
            pipe_length=request.pipe_length,
            pipe_type=request.pipe_type,
            shaft_diameter=request.shaft_diameter,
            bearing=request.bearing,
            bearing_make=request.bearing_make,
            housing=request.housing,
            weight_kg=request.weight_kg,
            unit_price=request.unit_price,
            rubber_diameter=request.rubber_diameter,
            belt_widths=request.belt_widths,
            quantity=request.quantity,
            shaft_end_type=request.shaft_end_type or "B",
            custom_shaft_extension=request.custom_shaft_extension
        )
        
        pdf_bytes = pdf_buffer.getvalue()
        filename = f"Drawing_{request.product_code.replace(' ', '_').replace('/', '-')}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate drawing: {str(e)}")


class EmailDrawingRequest(BaseModel):
    product_code: str
    roller_type: str
    pipe_diameter: float
    pipe_length: int
    pipe_type: str
    shaft_diameter: int
    bearing: str
    bearing_make: str
    housing: str
    weight_kg: float
    unit_price: float
    rubber_diameter: Optional[float] = None
    belt_widths: Optional[List[int]] = None
    quantity: int = 1
    shaft_end_type: Optional[str] = "B"
    custom_shaft_extension: Optional[int] = None
    recipient_email: str


@router.post("/email-drawing")
async def email_drawing(request: EmailDrawingRequest, current_user: dict = Depends(get_current_user)):
    """Generate a drawing PDF and email it to the recipient"""
    from drawing_generator import generate_roller_drawing
    
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        raise HTTPException(status_code=500, detail="Email service not configured")
    
    try:
        # Generate the PDF
        pdf_buffer = generate_roller_drawing(
            product_code=request.product_code,
            roller_type=request.roller_type,
            pipe_diameter=request.pipe_diameter,
            pipe_length=request.pipe_length,
            pipe_type=request.pipe_type,
            shaft_diameter=request.shaft_diameter,
            bearing=request.bearing,
            bearing_make=request.bearing_make,
            housing=request.housing,
            weight_kg=request.weight_kg,
            unit_price=request.unit_price,
            rubber_diameter=request.rubber_diameter,
            belt_widths=request.belt_widths,
            quantity=request.quantity,
            shaft_end_type=request.shaft_end_type or "B",
            custom_shaft_extension=request.custom_shaft_extension
        )
        
        pdf_bytes = pdf_buffer.getvalue()
        filename = f"Drawing_{request.product_code.replace(' ', '_').replace('/', '-')}.pdf"
        
        # Create email
        msg = MIMEMultipart()
        msg['From'] = GMAIL_USER
        msg['To'] = request.recipient_email
        msg['Subject'] = f"Roller Drawing - {request.product_code}"
        
        # Email body
        body = f"""
Dear Customer,

Please find attached the technical drawing for your requested roller:

Product Code: {request.product_code}
Roller Type: {request.roller_type}
Pipe Diameter: {request.pipe_diameter}mm
Pipe Length: {request.pipe_length}mm
Shaft Diameter: {request.shaft_diameter}mm
Bearing: {request.bearing}
Weight: {request.weight_kg}kg

For any queries, please contact us.

Best Regards,
Convero Solutions
        """
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach PDF
        pdf_attachment = MIMEApplication(pdf_bytes, _subtype='pdf')
        pdf_attachment.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(pdf_attachment)
        
        # Send email via Gmail SMTP
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.send_message(msg)
        
        return {"message": f"Drawing sent successfully to {request.recipient_email}"}
        
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=500, detail="Email authentication failed. Please check Gmail credentials.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")






# ============= USER MANAGEMENT (Admin only) =============

class UserRoleUpdate(BaseModel):
    email: str
    role: str


class UserCreate(BaseModel):
    email: str
    password: str = Field(..., min_length=6)
    name: str = Field(..., min_length=1)
    role: str
    company: Optional[str] = None
    designation: Optional[str] = None


@router.post("/admin/users")
async def create_user_as_admin(req: UserCreate, current_user: dict = Depends(get_current_user)):
    """Admin-only: create a new staff/customer user with a chosen role."""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    if req.role not in UserRole.assignable():
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of {UserRole.assignable()}")
    email = req.email.strip().lower()
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Invalid email address")
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    user_dict = {
        "email": email,
        "name": req.name.strip(),
        "role": req.role,
        "company": req.company,
        "designation": req.designation,
        "hashed_password": get_password_hash(req.password),
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user.get("email"),
    }
    if req.role == UserRole.CUSTOMER:
        user_dict["customer_code"] = await generate_customer_code()
    await db.users.insert_one(user_dict)
    return {
        "message": f"User {email} created with role {req.role}",
        "user": {
            "email": email,
            "name": user_dict["name"],
            "role": user_dict["role"],
            "company": user_dict.get("company"),
            "customer_code": user_dict.get("customer_code"),
        },
    }


@router.get("/admin/users")
async def list_users(current_user: dict = Depends(get_current_user)):
    """List all users (admin only)"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    users = await db.users.find(
        {},
        {"_id": 0, "password": 0, "push_token": 0}
    ).sort("created_at", -1).to_list(length=500)
    return {"users": users, "assignable_roles": UserRole.assignable()}


@router.put("/admin/users/role")
async def update_user_role(req: UserRoleUpdate, current_user: dict = Depends(get_current_user)):
    """Admin-only: change a user's role."""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    if req.role not in UserRole.assignable():
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of {UserRole.assignable()}")
    # Don't let admin demote themselves (prevent lockout)
    if req.email == current_user.get("email") and req.role != UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="You cannot change your own admin role")
    result = await db.users.update_one(
        {"email": req.email},
        {"$set": {"role": req.role, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": f"Role updated to {req.role} for {req.email}"}
